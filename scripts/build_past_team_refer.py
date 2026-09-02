"""
과거 조직 원본 엑셀에서 "1단계부서명"/"현소속부서명"/"비공식소속부서명" 3개
헤더 컬럼만 그대로 뽑아 <원본파일명>_team_refer.xlsx로 저장하는 스크립트
(2026-09-02, 사용자 확정 — 이 파일의 이전 버전을 완전히 대체).

이전 버전과의 차이: 이전 버전은 "End of Month Headcount" 162컬럼 고정 포맷을
전제로 BF/S/EV/ES 등 고정 열 위치 + VLOOKUP 방식 상위부서명 계산 + 월별
team_change.csv 비교까지 했다. 이번 버전은 그런 계산이 전혀 없다 — 헤더
텍스트로 3개 컬럼을 찾아 있는 그대로 복사하고, 중복 제거·정렬만 한다(원본
파일마다 열 위치가 달라도 헤더 텍스트만 맞으면 동작).

입력: data/raw/past_team_refer/ 안의 모든 .xlsx (Excel이 파일을 열어 두었을 때
생기는 임시 잠금 파일 "~$*.xlsx"는 제외). 각 파일은 2번째 행이 헤더이고 1번째
행은 무시한다. 시트는 항상 첫 번째 시트만 쓴다(사용자 확정).

출력: data/processed/past_team_refer/<원본파일명>_team_refer.xlsx
  - A열 = 1단계부서명(dep_name), B열 = 현소속부서명(middle_dep_name),
    C열 = 비공식소속부서명(pjt_part_name) — 헤더 텍스트는 원본 그대로 유지.
  - 헤더는 새 파일의 2행에 위치, 1행은 원본과 동일하게 비워둔다.
  - A~C 조합 기준 중복 제거(먼저 나온 행 유지), A~C가 전부 빈 행은 제외하되
    일부만 빈 행은 그대로 유지한다(사용자 확정).
  - 정렬: 안정 정렬을 C→B→A 순서로 연쇄 적용해, 최종적으로 A(1단계부서명)가
    1순위, B(현소속부서명)가 2순위, C(비공식소속부서명)가 3순위인 오름차순이
    되도록 한다(조직 위계상 큰 단위부터 정렬하는 게 자연스럽다는 해석 —
    사용자가 원래 "C→B→A 오름차순 정렬"이라고 표현한 것을 이 순서로 구현).

읽기: 사내 DRM이 걸린 xlsx는 xlwings(Excel COM)로만 열 수 있어
pipeline/excel_reader.read_xlsx()를 그대로 쓴다 — Windows PC에 Excel이
설치돼 있어야 하고, 없으면 자동으로 pandas 방식으로 폴백한다(그 경우 원본이
실제 DRM 파일이면 다시 같은 에러가 난다). 이 스크립트 자체는 xlwings/Excel이
없는 환경(리눅스 개발 서버 등)에서도 비DRM 파일로 로직 검증은 가능하다.

헤더 일부만 있는 파일 처리: 3개 헤더 중 하나라도 없으면 그 파일은 실패
처리하고 다음 파일로 넘어간다(부분 헤더로 어설프게 만들지 않음, 사용자
확정). 마지막에 총 파일 수/성공/실패 개수를 요약해 출력한다.

사용법:
  python scripts/build_past_team_refer.py
"""
import glob
import os
import sys

from openpyxl import Workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

# xlwings(실제 Excel COM 자동화) 기반 리더 — 이 프로젝트의 다른 모든 원천
# 처리기와 동일하게, 사내 DRM이 걸린 xlsx는 plain pandas/openpyxl로 못 열고
# ("Excel file format cannot be determined" 에러) xlwings로 실제 Excel을
# 띄워서 읽어야 한다(pipeline/excel_reader.py 독스트링 참고).
from pipeline.excel_reader import clean_str, read_xlsx  # noqa: E402

RAW_DIR = os.path.join(BASE_DIR, 'data', 'raw', 'past_team_refer')
OUT_DIR = os.path.join(BASE_DIR, 'data', 'processed', 'past_team_refer')

# (원본 헤더 텍스트, 의미상 영문 이름) — 순서 그대로 출력 A/B/C열이 된다.
HEADERS = [
    ('1단계부서명', 'dep_name'),
    ('현소속부서명', 'middle_dep_name'),
    ('비공식소속부서명', 'pjt_part_name'),
]
HEADER_TEXTS = [h for h, _ in HEADERS]


def _list_source_files() -> list:
    """RAW_DIR 안의 모든 .xlsx — Excel이 파일을 열어 둔 동안 생기는 임시
    잠금 파일("~$"로 시작)은 제외한다."""
    all_xlsx = glob.glob(os.path.join(RAW_DIR, '*.xlsx'))
    return sorted(p for p in all_xlsx if not os.path.basename(p).startswith('~$'))


def process_file(path: str) -> tuple:
    """한 원본 파일을 처리해 (성공 여부, 출력 경로 또는 None, 행 수, 에러 메시지) 반환.
    read_xlsx()의 header_row는 0-based이므로 1을 넘기면 물리적 2번째 행이
    헤더가 된다(1번째 행은 자동으로 무시)."""
    try:
        df = read_xlsx(path, sheet=0, header_row=1)
    except Exception as exc:
        return False, None, 0, f'파일 읽기 실패: {exc}'

    missing = [h for h in HEADER_TEXTS if h not in df.columns]
    if missing:
        return False, None, 0, f'헤더 없음: {", ".join(missing)}'

    # 3개 컬럼 값을 그대로 뽑는다(가공 없음) — clean_str()으로 None/NaN류만
    # 빈 문자열로 통일(excel_reader.py 공통 관례).
    rows = []
    seen = set()
    for _, row in df[HEADER_TEXTS].iterrows():
        vals = tuple(clean_str(row[h]) for h in HEADER_TEXTS)
        if not any(vals):          # A~C 전부 빈 행은 제외(일부만 빈 행은 유지)
            continue
        if vals in seen:           # A+B+C 조합 중복 제거(먼저 나온 행 유지)
            continue
        seen.add(vals)
        rows.append(vals)

    # 안정 정렬을 C→B→A 순으로 연쇄 적용 → 최종 A 1순위/B 2순위/C 3순위 오름차순.
    rows.sort(key=lambda r: r[2])
    rows.sort(key=lambda r: r[1])
    rows.sort(key=lambda r: r[0])

    os.makedirs(OUT_DIR, exist_ok=True)
    base = os.path.splitext(os.path.basename(path))[0]
    out_path = os.path.join(OUT_DIR, f'{base}_team_refer.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = 'team_refer'
    # 1행은 원본과 동일하게 비워두고, 헤더는 2행에 원본 텍스트 그대로.
    for col_idx, header in enumerate(HEADER_TEXTS, start=1):
        ws.cell(row=2, column=col_idx, value=header)
    for r_idx, vals in enumerate(rows, start=3):
        for col_idx, v in enumerate(vals, start=1):
            ws.cell(row=r_idx, column=col_idx, value=v)
    wb.save(out_path)

    return True, out_path, len(rows), None


def main():
    if not os.path.isdir(RAW_DIR) or not _list_source_files():
        os.makedirs(RAW_DIR, exist_ok=True)
        print(f'[안내] {RAW_DIR} 에 원본 엑셀 파일을 먼저 넣어주세요.')
        return

    files = _list_source_files()
    success = 0
    fail = 0
    for path in files:
        name = os.path.basename(path)
        ok, out_path, n_rows, err = process_file(path)
        if ok:
            success += 1
            print(f'[OK]   {name} → {os.path.basename(out_path)} ({n_rows}행)')
        else:
            fail += 1
            print(f'[실패] {name}: {err}', file=sys.stderr)

    print(f'\n총 {len(files)}개 중 성공 {success}개, 실패 {fail}개')


if __name__ == '__main__':
    main()
