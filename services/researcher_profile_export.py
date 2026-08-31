"""
"보유 전문성" 자연어 질문 결과에서 선택한 연구원들의 프로필을 엑셀로 내보낸다
(`pages/researcher_similarity_map.py`의 "엑셀 다운로드" 버튼 → 모달에서 사용).

컬럼 구성/표기 규칙은 사용자와의 인터뷰로 확정된 것을 그대로 코드화한 것 —
바꾸려면 아래 `_COLUMNS`와 각 `_col_*` 함수만 수정하면 된다. 규칙 요약:
  - 사번: researcher_id 8자리 0패딩(data_store.read_processed가 이미 패딩해 줌)
  - 값이 없으면 전부 "-" (다중 이력 필드는 이력이 하나도 없을 때 셀 전체가 "-")
  - 날짜는 대부분 'YY(2자리 연도, 예: '24)로 축약 표기 — 원본 협의 그대로
  - 학력/과제수행이력/양성이력/핵심이력은 한 셀 안에 줄바꿈(\\n)으로 여러 줄
"""
import io
import math
from datetime import date, datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from components.timeline_data import dedupe_patents, job_points
from services import auth, data_store, evaluations
from services import language_qualification as language_qual
from services import work_experience as work_exp
from services.task_history import merge_task_rows, task_year
from services.task_history import sort_key as _task_sort_key

_FONT_NAME = '바탕체'
_FONT_SIZE = 11
_THIN = Side(style='thin', color='000000')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# 평가('24~'26) 컬럼 — 회계연도(매년 3월 시작, services.evaluations 참고) 기준
# 최근 3개년. 헤더 문자열은 모듈 임포트 시점에 한 번 계산(연 1회만 바뀌므로
# 요청마다 다시 계산할 필요 없음 — 다른 회계연도 경계를 넘기려면 프로세스
# 재시작만 있으면 됨, 이 앱의 다른 "현재 시점 기준" 값들과 동일한 전제).
# evaluation_years()는 최신 연도가 먼저 오는 내림차순([2026,2025,2024])을
# 반환하는데, 헤더("'24~'26")는 오름차순으로 읽히므로 오름차순으로 정렬해야
# 셀 값 순서와 헤더가 맞는다(2026-08-29 발견·수정 — pages/researcher_list.py는
# 이미 sorted()로 오름차순을 쓰고 있었는데 이 모듈만 정렬 없이 그대로 썼다).
_EVAL_SALARY_YEARS = sorted(evaluations.evaluation_years()[0])
_EVAL_HALF_YEARS = sorted(evaluations.evaluation_years()[1])
_EVAL_HEADER = f"평가\n('{str(_EVAL_SALARY_YEARS[0])[-2:]}~'{str(_EVAL_SALARY_YEARS[-1])[-2:]})"

_DEGREE_ORDER = ['박사', '석사', '학사']
_DEGREE_CODE = {'박사': '박', '석사': '석', '학사': '학'}

_PROMOTION_REF_BASE = date(2027, 3, 1)


def _s(v) -> str:
    """빈 값/NaN을 빈 문자열로 통일."""
    if v is None:
        return ''
    s = str(v).strip()
    return '' if s.lower() in ('', 'nan', 'none', 'nat') else s


def _or_dash(v) -> str:
    s = _s(v)
    return s if s else '-'


def _birth_year_int(v) -> int | None:
    """researchers.csv의 birth_year를 정수로 안전하게 파싱한다.
    data_store.read_processed()가 researcher_id 외 컬럼은 dtype을 지정하지
    않고 CSV를 읽는데, birth_year가 하나라도 비어있는(NaN) 행이 있으면
    pandas가 컬럼 전체를 float로 추론해 정상 값도 "1990.0"처럼 소수점이
    붙어 들어온다(2026-08-29 발견 — 엑셀 다운로드/AI 검색 결과의 나이가
    전부 "-"로 나오던 원인). 단순 `.isdigit()` 체크는 "1990.0"을 숫자로
    인정하지 않아 실패하므로, float로 한 번 변환한 뒤 int로 반올림 없이
    잘라 정수를 얻는다."""
    s = _s(v)
    if not s:
        return None
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return None


def _yy(date_str) -> str:
    s = _s(date_str)
    if len(s) >= 4 and s[:4].isdigit():
        return "'" + s[2:4]
    return ''




def _next_promotion_ref_date(today: date) -> date:
    """직급/년차 기준일 — 2027-03-01부터 시작해 오늘을 지나지 않은 첫 3/1."""
    ref = _PROMOTION_REF_BASE
    while ref < today:
        ref = ref.replace(year=ref.year + 1)
    return ref


def _floor1(x: float) -> float:
    return math.floor(x * 10) / 10


def position_years(promotion_date: str) -> int | None:
    """"CL/년차"의 "년차" 부분만 계산해 정수로 반환 — promotion_date가 없으면
    계산 불가라 None(호출부가 "직급 표시만" 등으로 폴백). _col_position_year()
    뿐 아니라 pipeline/process_researcher_similarity.py의 CL 기반 시니어/주니어
    분류에서도 재사용한다(같은 "몇 년차인지" 계산을 두 곳에서 따로 하면
    어긋날 위험이 있어 여기 하나로 모음)."""
    promo = _s(promotion_date)
    if not promo:
        return None
    promo_dt = date.fromisoformat(promo[:10])
    ref = _next_promotion_ref_date(date.today())
    return int(_floor1((ref - promo_dt).days / 365))


def _load_tables() -> dict:
    return {
        'researchers': data_store.read_processed('researchers'),
        'education': data_store.read_processed('education'),
        'evaluations': data_store.read_processed('evaluations'),
        'tasks': data_store.read_processed('tasks'),
        'nurturing': data_store.read_processed('nurturing'),
        'incentive_selection': data_store.read_processed('incentive_selection'),
        'team_refer': data_store.read_processed('team_refer'),
        'tech_ownership': data_store.read_processed('tech_ownership'),
        'patents': data_store.read_processed('patents'),
        'publications': data_store.read_processed('publications'),
        'job_profile': data_store.read_processed('job_profile'),
        'language_qualification': data_store.read_processed('language_qualification'),
        'work_experience': data_store.read_processed('work_experience'),
        'expertise_profiles': data_store.read_expertise_profiles(),
    }


def _rows_for(df, researcher_id: str):
    if df is None or df.empty or 'researcher_id' not in df.columns:
        return []
    return df[df['researcher_id'] == researcher_id].to_dict('records')


def _df_for(df, researcher_id: str):
    """_rows_for()의 DataFrame 버전 — dedupe_patents()처럼 DataFrame을 그대로
    받아야 하는 헬퍼(특허 국가별 중복 합치기)에 넘길 때 사용."""
    if df is None or df.empty or 'researcher_id' not in df.columns:
        return df.iloc[0:0] if df is not None else pd.DataFrame()
    return df[df['researcher_id'] == researcher_id]


def _col_id(rid, _rows):
    return rid


def _col_knox(_rid, rows):
    return _or_dash(rows['researcher'].get('knox_id') if rows['researcher'] else None)


def _col_name_gender_age(_rid, rows):
    r = rows['researcher']
    if not r:
        return '-'
    name = _or_dash(r.get('name'))
    gender = _s(r.get('gender')) or '-'
    birth_year = _birth_year_int(r.get('birth_year'))
    age = f'{datetime.now().year - birth_year}세' if birth_year is not None else '-'
    return f'{name}\n({gender}/{age})'


def _col_dept_task(_rid, rows):
    r = rows['researcher']
    if not r:
        return '-'
    org_code = str(r.get('org_code') or '').strip()
    dep_map, pjt_map = rows.get('dep_pjt_maps') or ({}, {})
    # 부서/과제(파트) = team_refer의 dep_name/pjt_part_name(연구원 명단 표와
    # 동일 기준 — 사용자 확정). 매핑이 없으면 원본 department/org_code
    # 그대로 보여준다(빈 칸이면 데이터 누락처럼 보이므로).
    dept = _or_dash(dep_map.get(org_code) or r.get('department'))
    org = _or_dash(pjt_map.get(org_code) or org_code)
    return f'{dept}\n({org})'


def _col_hire_date(_rid, rows):
    r = rows['researcher']
    hire_date = _s(r.get('hire_date')) if r else ''
    if not hire_date:
        return '-'
    hd = date.fromisoformat(hire_date[:10])
    tenure = round((date.today() - hd).days / 365, 1)
    return f"{hd.strftime('%Y%m%d')}\n({tenure}년)"


def _col_education(_rid, rows):
    by_degree = {}
    for e in rows['education']:
        deg = _s(e.get('degree'))
        if deg in _DEGREE_CODE:
            by_degree.setdefault(deg, e)
    lines = []
    for deg in _DEGREE_ORDER:
        e = by_degree.get(deg)
        if not e:
            continue
        code = _DEGREE_CODE[deg]
        school = _or_dash(e.get('school'))
        major = _or_dash(e.get('major'))
        lines.append(f'{code}){school} {major}')
    return '\n'.join(lines) if lines else '-'


def _col_evaluation(_rid, rows):
    """평가 — 연봉등급 3개년을 첫 줄에 헤더("'24~'26")와 같은 오름차순으로
    "다/다/다", 그에 대응하는(각 연봉등급 연도 - 1) 역량/하반기업적 3개년을
    둘째 줄에 역시 오름차순으로 "(VG/MT, VG/MT, VG/MT)"로 표시(2026-08-29
    수정 — _EVAL_SALARY_YEARS/_EVAL_HALF_YEARS를 정렬 없이 그대로 써서
    evaluation_years()의 내림차순(최신 연도가 먼저)이 그대로 노출돼 헤더와
    셀 값 순서가 어긋나 있었다). 둘째 줄의 각 항목은
    evaluations.format_half_display()로 만드는데,
    그 해 연봉등급이 있으면 역량/하반기업적 중 있는 것만 보여주고(예: 역량
    없음/하반기만 있음 → "MT"만, "-/MT"처럼 빈 자리를 표시하지 않음 —
    2026-08-29 확정, 예전엔 상반기업적과 짝지었으나 역량으로 교체), 연봉등급
    자체가 없으면 기존대로 상/하반기업적 두 자리를 항상 표시한다(빈 자리는
    '-').

    화면(pages/*.py)이 view_evaluation 권한 없는 역할에는 평가등급을 아예
    안 보여주는 것과 동일한 기준을 여기 엑셀 다운로드에도 적용한다 — 권한
    없으면 평가 데이터를 셀에 채우지 않는다(권한은 build_profile_workbook()이
    행마다 다시 확인하지 않도록 rows['permissions']로 한 번만 계산해 전달).

    view_evaluation이 있어도 부서 단위 평가 제외(2026-08-31, rows의
    'eval_excluded_dep_ids'/'org_code_dep_id_map')에 해당하는 연구원이면
    화면(researcher_profile.py/researcher_list.py)과 동일하게 이 행만 가린다."""
    if not rows['permissions']['view_evaluation']:
        return '-'
    excluded_dep_ids = rows.get('eval_excluded_dep_ids')
    if excluded_dep_ids:
        researcher = rows.get('researcher') or {}
        org_code = _s(researcher.get('org_code'))
        dep_id = (rows.get('org_code_dep_id_map') or {}).get(org_code, '')
        if dep_id in excluded_dep_ids:
            return '-'
    eva_rows = rows['evaluations']
    eva = eva_rows[0] if eva_rows else {}
    salary_line = '/'.join(_s(eva.get(evaluations.salary_grade_column(y))) or '-' for y in _EVAL_SALARY_YEARS)
    half_line = '(' + ', '.join(
        evaluations.format_half_display(
            _s(eva.get(evaluations.salary_grade_column(y + 1))),
            _s(eva.get(evaluations.first_half_column(y))),
            _s(eva.get(evaluations.second_half_column(y))),
            _s(eva.get(evaluations.competency_column(y))),
        )
        for y in _EVAL_HALF_YEARS
    ) + ')'
    return f'{salary_line}\n{half_line}'


def _col_position_year(_rid, rows):
    r = rows['researcher']
    if not r:
        return '-'
    position = _or_dash(r.get('position'))
    years = position_years(r.get('promotion_date'))
    if years is None:
        return position
    return f'{position}-{years}'


def _col_position_title(_rid, rows):
    """직책 = team_refer.csv의 assignment_name(조직장급만 등록돼 있어 대부분
    매핑이 없음 — 그 경우 "-")."""
    team_rows = rows['team_refer']
    return _or_dash(team_rows[0].get('assignment_name')) if team_rows else '-'


def _col_tasks(_rid, rows):
    """과제수행이력 — 같은 과제명(the_task_name 우선, 없으면 task_name)의
    여러 참여 구간을 services.task_history.merge_task_rows()가 연도 기준
    포함/연결 규칙으로 하나로 합친 뒤(화면 표 tasks_block()과 공유하는
    로직 — 사용자 요청 2026-08-31 "화면에도 동일하게 반영"), 진행중('현재')
    과제를 맨 위로, 그 다음은 최근 시작연도 순으로 정렬한다."""
    merged = merge_task_rows(rows['tasks'])
    merged.sort(key=_task_sort_key)
    lines = []
    for m in merged:
        start_year = task_year(m['start_row'].get('start_date'))
        start = f"'{start_year % 100:02d}" if start_year is not None else '-'
        if m['current']:
            end = '현재'
        else:
            end_year = task_year(m['end_row'].get('end_date')) if m['end_row'] else None
            end = f"'{end_year % 100:02d}" if end_year is not None else '-'
        lines.append(f"{m['name']}({start} ~ {end})")
    return '\n'.join(lines) if lines else '-'


def _col_nurturing(_rid, rows):
    items = sorted(rows['nurturing'], key=lambda n: _s(n.get('start_date')))
    lines = []
    for n in items:
        start = _yy(n.get('start_date')) or '-'
        end = _yy(n.get('end_date')) or '-'
        subcategory = _or_dash(n.get('subcategory'))
        institution = _or_dash(n.get('institution'))
        lines.append(f"{start}~{end} / {subcategory} / {institution}")
    return '\n'.join(lines) if lines else '-'


def _col_incentive(_rid, rows):
    """핵심이력 — incentive_selection.csv 기반. 화면이 view_incentive 권한
    없는 역할에는 이 데이터를 안 보여주는 것과 동일하게 여기서도 가린다
    (권한은 rows['permissions']로 한 번만 계산해 전달받음 — _col_evaluation 참고)."""
    if not rows['permissions']['view_incentive']:
        return '-'
    items = [i for i in rows['incentive_selection'] if _s(i.get('selected')).lower() in ('true', '1')]
    items.sort(key=lambda i: _s(i.get('year')))
    lines = [f"{_yy(i.get('year'))} : {_or_dash(i.get('category'))}" for i in items]
    return '\n'.join(lines) if lines else '-'


def _col_tech_ownership(_rid, rows):
    """보유기술 — components/detail_tabs.py의 _tech_ownership_table()과 동일하게
    tech_ownership.csv의 tech_1~5/lv_1~5/portion_1~5를 "전문분야(Lv N, 보유율 M%)"
    형태로 한 셀에 줄바꿈 나열한다(화면의 '보유기술' 표와 동일 항목, 데이터 없는
    슬롯은 건너뜀)."""
    tech_rows = rows['tech_ownership']
    if not tech_rows:
        return '-'
    tech_row = tech_rows[0]
    lines = []
    for i in range(1, 6):
        name = _s(tech_row.get(f'tech_{i}'))
        if not name:
            continue
        lv = _s(tech_row.get(f'lv_{i}')) or '-'
        portion = _s(tech_row.get(f'portion_{i}'))
        portion_disp = f'{portion}%' if portion else '-'
        lines.append(f'{name} (Lv {lv}, 보유율 {portion_disp})')
    return '\n'.join(lines) if lines else '-'


def expertise_field_lines(profile: dict | None, field: str) -> str:
    """보유 전문성(LLM) 프로필 dict에서 필드 하나(strength_fields 등)를 줄바꿈
    나열 문자열로 뽑는다. services/similarity_map.py의 조직도 기반 다운로드처럼
    _researcher_row_context() 없이 profile dict만 있는 호출부에서도 같은 서식을
    쓸 수 있도록 _expertise_field()의 내부 로직을 공개 함수로 분리했다."""
    items = (profile.get(field) if profile else None) or []
    return '\n'.join(items) if items else '-'


def _expertise_field(field: str):
    """보유 전문성(LLM)의 한 필드를 한 셀에 줄바꿈으로 나열하는 컬럼 함수를 만든다.
    components/detail_tabs.py의 llm_summary_block()과 동일한 4개 필드
    (strength_fields/strength_keywords/key_responsibilities/domain_knowledge_skill)를
    강점 분야/강점 키워드/주요 역할·책임/전문지식 및 역량 4개 컬럼으로 나눠 받고
    싶다는 요청에 따라 컬럼별 함수로 분리했다. 다운로드 시 선택 사항(옵트인)."""
    def _col(_rid, rows):
        return expertise_field_lines(rows.get('expertise_profile'), field)
    return _col


_EXPERTISE_COLUMNS = [
    ('보유전문성(강점 분야)', _expertise_field('strength_fields')),
    ('보유전문성(강점 키워드)', _expertise_field('strength_keywords')),
    ('보유전문성(주요 역할·책임)', _expertise_field('key_responsibilities')),
    ('보유전문성(전문지식 및 역량)', _expertise_field('domain_knowledge_skill')),
]


def _col_patents(_rid, rows):
    """특허 실적 — components/detail_tabs.py의 patents_tab()과 동일하게
    dedupe_patents()로 국가별 중복 출원 행을 하나로 합친 뒤, 출원일 내림차순으로
    "출원일 : 발명명칭 (상태, 대표발명자, 지분율%, 등급)"를 한 셀에 줄바꿈 나열."""
    pat = rows.get('patents_df')
    if pat is None or pat.empty:
        return '-'
    pat_dedup = dedupe_patents(pat)
    sort_col = 'application_date' if 'application_date' in pat_dedup.columns else pat_dedup.columns[0]
    lines = []
    for _, p in pat_dedup.sort_values(sort_col, ascending=False).iterrows():
        date = _s(p.get('application_date'))[:7] or '-'
        title = _s(p.get('title')) or _s(p.get('title_ko')) or '-'
        status = _s(p.get('status')) or '-'
        lead = '대표' if _s(p.get('is_lead_inventor')).lower() in ('y', '1', 'true') else ''
        share = _s(p.get('share_ratio'))
        share_disp = f'{share}%' if share else ''
        grade = _s(p.get('patent_grade'))
        grade_a = _s(p.get('patent_grade_a_sub'))
        grade_disp = grade + (f'({grade_a})' if grade_a else '') if grade else ''
        extras = ', '.join(v for v in (status, lead, share_disp, grade_disp) if v)
        lines.append(f'{date} : {title}' + (f' ({extras})' if extras else ''))
    return '\n'.join(lines) if lines else '-'


def _col_publications(_rid, rows):
    """논문 실적 — components/detail_tabs.py의 publications_tab()과 동일하게
    pub_date(없으면 pub_year) 내림차순으로 "발표일 : 제목 (게재처, 순위/총수,
    기여도%, 교신)"를 한 셀에 줄바꿈 나열."""
    items = rows.get('publications') or []
    if not items:
        return '-'
    items = sorted(items, key=lambda p: _s(p.get('pub_date')) or _s(p.get('pub_year')), reverse=True)
    lines = []
    for p in items:
        date = (_s(p.get('pub_date')) or _s(p.get('pub_year')))[:7] or '-'
        title = _s(p.get('title')) or '-'
        journal = _s(p.get('journal'))
        rank, total = _s(p.get('author_rank')), _s(p.get('total_authors'))
        rank_total = f'{rank}/{total}' if rank and total else ''
        contrib = _s(p.get('contribution'))
        contrib_disp = f'기여도 {contrib}%' if contrib else ''
        corr = '교신' if _s(p.get('is_corresponding')).lower() in ('true', '1', 'y', 'yes') else ''
        extras = ', '.join(v for v in (journal, rank_total, contrib_disp, corr) if v)
        lines.append(f'{date} : {title}' + (f' ({extras})' if extras else ''))
    return '\n'.join(lines) if lines else '-'


def _col_job_function(_rid, rows):
    """직무 — researchers.csv의 job_function(현재 시점 직무 분류) 값 그대로."""
    r = rows.get('researcher')
    return _or_dash(r.get('job_function') if r else None)


def _col_employment_status(_rid, rows):
    """재직상태 — researchers.csv의 employment_status(휴직/재직/퇴직 등,
    pipeline/process_researchers.py 원본 그대로) 값. is_current(최신
    인력현황 파일 포함 여부)와는 별개 개념이라 값이 다를 수 있다."""
    r = rows.get('researcher')
    return _or_dash(r.get('employment_status') if r else None)


def _col_job_profile(_rid, rows):
    """직무이력 — job_profile.csv(wide: job_profile_name_i/job_start_date_i/
    job_end_date_i)를 components/timeline_data.job_points()로 슬롯 해제한 뒤,
    시작일 내림차순으로 "직무명(YY ~ YY/현재)"를 한 셀에 줄바꿈 나열
    (과제수행이력 표기 규칙과 동일)."""
    points = job_points(rows.get('job_profile_df', pd.DataFrame()))
    if not points:
        return '-'
    points.sort(key=lambda p: p['start'], reverse=True)
    lines = []
    for p in points:
        start = p['start'].strftime("'%y") if p['start'] is not None else '-'
        end = p['end'].strftime("'%y") if p['end'] is not None else '현재'
        lines.append(f"{p['name']}({start} ~ {end})")
    return '\n'.join(lines)


def _col_language(_rid, rows):
    """어학 — language_qualification.csv(언어별 1행)를 services.
    language_qualification.format_lines()로 "{언어} {등급}(만료일 {날짜})"
    줄로 만들어 한 셀에 줄바꿈 나열(사용자 확정 — 프로필 화면과 동일한
    형식·헬퍼 공유). 보유 언어가 없으면 '-'."""
    lines = language_qual.format_lines(rows.get('language_qualification') or [])
    return '\n'.join(lines) if lines else '-'


def _col_work_experience(_rid, rows):
    """근무 경력 — work_experience.csv(회사별 1행)를 services.work_experience.
    format_lines()로 "회사명(시작'YY.MM ~ 종료'YY.MM, 직무명)" 줄로 만들어
    한 셀에 줄바꿈 나열(사용자 확정 — 프로필 화면과 동일한 형식·헬퍼
    공유). 경력이 없으면 '-'."""
    lines = work_exp.format_lines(rows.get('work_experience') or [])
    return '\n'.join(lines) if lines else '-'


_PATENT_COLUMNS = [('특허 실적', _col_patents)]
_PUBLICATION_COLUMNS = [('논문 실적', _col_publications)]
_JOB_FUNCTION_COLUMNS = [('직무', _col_job_function)]
_JOB_PROFILE_COLUMNS = [('직무이력', _col_job_profile)]
_EMPLOYMENT_STATUS_COLUMNS = [('재직상태', _col_employment_status)]
_LANGUAGE_COLUMNS = [('어학', _col_language)]
_WORK_EXPERIENCE_COLUMNS = [('근무 경력', _col_work_experience)]


# (헤더, 값 계산 함수) — 순서 = 엑셀 컬럼 순서
_COLUMNS = [
    ('사번', _col_id),
    ('Knox ID', _col_knox),
    ('성명\n(성별/나이)', _col_name_gender_age),
    ('부서\n(과제)', _col_dept_task),
    ('입사일', _col_hire_date),
    ('학력', _col_education),
    (_EVAL_HEADER, _col_evaluation),
    ('CL/년차', _col_position_year),
    ('직책', _col_position_title),
    ('과제수행이력', _col_tasks),
    ('양성이력', _col_nurturing),
    ('핵심이력', _col_incentive),
    ('보유기술', _col_tech_ownership),
]


def _researcher_row_context(researcher_id: str, tables: dict, permissions: dict,
                             dep_pjt_maps: tuple | None = None,
                             eval_excluded_dep_ids: set | None = None,
                             org_code_dep_id_map: dict | None = None) -> dict:
    researcher_rows = _rows_for(tables['researchers'], researcher_id)
    return {
        'dep_pjt_maps': dep_pjt_maps,
        # 부서 단위 평가 제외(2026-08-31) — _col_evaluation 참고. 요청 단위로
        # 한 번만 계산해 전달받는다(dep_pjt_maps와 같은 이유).
        'eval_excluded_dep_ids': eval_excluded_dep_ids,
        'org_code_dep_id_map': org_code_dep_id_map,
        'researcher': researcher_rows[0] if researcher_rows else None,
        'education': _rows_for(tables['education'], researcher_id),
        'evaluations': _rows_for(tables['evaluations'], researcher_id),
        'tasks': _rows_for(tables['tasks'], researcher_id),
        'nurturing': _rows_for(tables['nurturing'], researcher_id),
        'incentive_selection': _rows_for(tables['incentive_selection'], researcher_id),
        'team_refer': _rows_for(tables['team_refer'], researcher_id),
        'tech_ownership': _rows_for(tables['tech_ownership'], researcher_id),
        'patents_df': _df_for(tables['patents'], researcher_id),
        'publications': _rows_for(tables['publications'], researcher_id),
        'job_profile_df': _df_for(tables['job_profile'], researcher_id),
        'language_qualification': _rows_for(tables['language_qualification'], researcher_id),
        'work_experience': _rows_for(tables['work_experience'], researcher_id),
        'expertise_profile': tables['expertise_profiles'].get(researcher_id),
        # 요청(로그인 사용자) 단위로 한 번만 계산해 매 행마다 auth.can()을
        # 다시 호출하지 않도록 build_profile_workbook()에서 전달받는다
        # (_col_evaluation/_col_incentive 참고).
        'permissions': permissions,
    }


def researcher_name_map() -> dict:
    """researcher_id -> name 매핑(엑셀 다운로드 후보 라벨 생성용)."""
    df = data_store.read_processed('researchers')
    if df.empty or 'name' not in df.columns:
        return {}
    return dict(zip(df['researcher_id'], df['name']))


def candidate_label(researcher_id: str, name_map: dict | None = None) -> str:
    """모달 체크리스트에 보여줄 "이름 (사번)" 라벨. 이름 없으면 사번만."""
    name_map = name_map if name_map is not None else researcher_name_map()
    name = _s(name_map.get(researcher_id))
    return f'{name} ({researcher_id})' if name else researcher_id


# 자연어 질문(nl_query/open_data_query) 결과가 "사람에 대한 데이터"로 판단되면
# (결과에 researcher_id 컬럼이 있으면) 항상 맨 앞에 붙이는 7개 기본 컬럼.
# 값 계산 로직은 엑셀 다운로드와 동일한 걸 재사용(같은 사람은 어디서 봐도
# 같은 표기) — 학력은 엑셀처럼 전체 이력이 아니라 "최종 학력 1건만", CL/년차는
# _col_position_year()를 그대로 재사용(예: "CL4-17", 승격기준일 없으면 "CL4").
PERSON_BASE_COLUMNS = ['researcher_id', 'name', 'department', 'org_code', 'position_year', 'degree_major', 'age']

# _highest_degree_str() 전용 — 엑셀 다운로드의 _col_education()(박/석/학사만,
# 나머지 제외)과 달리 여기서는 전공만 필터에서 뺄 뿐 학력 자체는 전문대/고교
# 까지 전부 인정한다(process_education.py의 DEG_ORDER와 동일한 5단계 우선순위
# — education.csv 자체가 이미 "학사 이상이 있으면 전문대/고교 제외" 규칙으로
# 정리돼 있어서, 여기 남아 있는 전문대/고교는 그게 그 사람의 최종 학력이라는 뜻).
_DEGREE_ORDER_FULL = ['박사', '석사', '학사', '전문대', '고교']
_DEGREE_CODE_FULL = {'박사': '박', '석사': '석', '학사': '학', '전문대': '전', '고교': '고'}


def highest_degree_row(education_rows: list) -> dict | None:
    """education.csv 행 목록(한 사람 것) 중 최종학력(박사>석사>학사>전문대>고교
    우선순위, process_education.py의 DEG_ORDER와 동일) 1건을 반환. 없으면 None.
    services.nl_query의 학력 조건 검색(find_researchers_by_criteria)과
    _highest_degree_str() 둘 다 이 함수를 공유한다."""
    by_degree = {}
    for e in education_rows:
        deg = _s(e.get('degree'))
        if deg in _DEGREE_CODE_FULL:
            by_degree.setdefault(deg, e)
    for deg in _DEGREE_ORDER_FULL:
        e = by_degree.get(deg)
        if e:
            return e
    return None


def _highest_degree_str(education_rows: list) -> str:
    e = highest_degree_row(education_rows)
    if not e:
        return '-'
    code = _DEGREE_CODE_FULL[_s(e.get('degree'))]
    school = _or_dash(e.get('school'))
    major = _or_dash(e.get('major'))
    return f'{code}){school} {major}'


def person_base_table(researcher_ids: list) -> dict:
    """PERSON_BASE_COLUMNS 순서에 맞는 값 리스트를 researcher_id별로 반환.
    researchers.csv/education.csv를 한 번만 읽어 여러 명을 한꺼번에 처리한다
    (nl_query/open_data_query가 결과 테이블 렌더링 때마다 호출하므로 매번
    파일을 다시 읽지 않도록)."""
    researchers_df = data_store.read_processed('researchers')
    education_df = data_store.read_processed('education')
    researchers_by_id = (
        researchers_df.set_index('researcher_id').to_dict('index') if not researchers_df.empty else {}
    )
    current_year = datetime.now().year

    out = {}
    for rid in researcher_ids:
        r = researchers_by_id.get(rid)
        edu_rows = _rows_for(education_df, rid)
        degree_major = _highest_degree_str(edu_rows)
        if r:
            name = _or_dash(r.get('name'))
            department = _or_dash(r.get('department'))
            org_code = _or_dash(r.get('org_code'))
            position_year = _col_position_year(rid, {'researcher': r})
            birth_year = _birth_year_int(r.get('birth_year'))
            age = str(current_year - birth_year) if birth_year is not None else '-'
        else:
            name = department = org_code = position_year = age = '-'
        out[rid] = [rid, name, department, org_code, position_year, degree_major, age]
    return out


_COLUMN_WIDTHS = [12, 14, 16, 18, 12, 26, 22, 12, 10, 34, 30, 22, 30]
_EMPLOYMENT_STATUS_COLUMN_WIDTH = 12
_EXPERTISE_COLUMN_WIDTH = 26
_PATENT_COLUMN_WIDTH = 40
_PUBLICATION_COLUMN_WIDTH = 40
_JOB_FUNCTION_COLUMN_WIDTH = 14
_JOB_PROFILE_COLUMN_WIDTH = 30
_LANGUAGE_COLUMN_WIDTH = 26
_WORK_EXPERIENCE_COLUMN_WIDTH = 30


def build_profile_workbook(
    researcher_ids: list,
    include_expertise: bool = False,
    include_patents: bool = False,
    include_publications: bool = False,
    include_job_function: bool = False,
    include_job_profile: bool = False,
    include_employment_status: bool = False,
    include_language: bool = False,
    include_work_experience: bool = False,
) -> bytes:
    """선택된 researcher_id 목록으로 엑셀(xlsx) 바이트를 만들어 반환한다.
    양식: 바탕체 11pt, 전체 검정 테두리, 헤더만 볼드, 줄바꿈 셀은 자동 줄바꿈.
    include_*가 True인 항목만 해당 옵트인 컬럼 그룹(_PATENT_COLUMNS/
    _PUBLICATION_COLUMNS/_JOB_FUNCTION_COLUMNS/_JOB_PROFILE_COLUMNS/
    _EMPLOYMENT_STATUS_COLUMNS/_LANGUAGE_COLUMNS/_WORK_EXPERIENCE_COLUMNS/
    _EXPERTISE_COLUMNS)을 이 순서대로 맨 끝에
    추가한다 — 전부 기본값은 False(다운로드 화면 체크박스 기본 해제)이고,
    켜져도 _COLUMNS 자체는 건드리지 않고 이 함수 안에서만 로컬 사본에 덧붙인다."""
    tables = _load_tables()
    # 로그인 사용자당 한 번만 계산 — _col_evaluation/_col_incentive가 매 행마다
    # auth.can()을 다시 호출하지 않도록 _researcher_row_context()에 실어 보낸다.
    permissions = {'view_evaluation': auth.can('view_evaluation'), 'view_incentive': auth.can('view_incentive')}
    # org_code → dep_name/pjt_part_name 매핑도 배치당 한 번만 만든다(연구원
    # 수만큼 team_refer를 반복 스캔하지 않도록) — similarity_map이 이 모듈을
    # 임포트하므로(순환 임포트 방지) 여기서는 지연 임포트로 가져온다.
    from services import similarity_map
    dep_pjt_maps = similarity_map.org_code_label_maps()
    # 부서 단위 평가 제외(2026-08-31) — view_evaluation은 있어도 특정 부서
    # 소속 연구원의 평가만 가리는 사용자를 위해 화면(researcher_list.py/
    # researcher_profile.py)과 동일한 기준을 엑셀에도 적용한다.
    eval_excluded_dep_ids = auth.eval_excluded_dep_ids()
    org_code_dep_id_map = similarity_map.org_code_dep_id_map() if eval_excluded_dep_ids else {}

    columns = list(_COLUMNS)
    widths = list(_COLUMN_WIDTHS)
    if include_patents:
        columns.extend(_PATENT_COLUMNS)
        widths.extend([_PATENT_COLUMN_WIDTH] * len(_PATENT_COLUMNS))
    if include_publications:
        columns.extend(_PUBLICATION_COLUMNS)
        widths.extend([_PUBLICATION_COLUMN_WIDTH] * len(_PUBLICATION_COLUMNS))
    if include_job_function:
        columns.extend(_JOB_FUNCTION_COLUMNS)
        widths.extend([_JOB_FUNCTION_COLUMN_WIDTH] * len(_JOB_FUNCTION_COLUMNS))
    if include_job_profile:
        columns.extend(_JOB_PROFILE_COLUMNS)
        widths.extend([_JOB_PROFILE_COLUMN_WIDTH] * len(_JOB_PROFILE_COLUMNS))
    if include_employment_status:
        columns.extend(_EMPLOYMENT_STATUS_COLUMNS)
        widths.extend([_EMPLOYMENT_STATUS_COLUMN_WIDTH] * len(_EMPLOYMENT_STATUS_COLUMNS))
    if include_language:
        columns.extend(_LANGUAGE_COLUMNS)
        widths.extend([_LANGUAGE_COLUMN_WIDTH] * len(_LANGUAGE_COLUMNS))
    if include_work_experience:
        columns.extend(_WORK_EXPERIENCE_COLUMNS)
        widths.extend([_WORK_EXPERIENCE_COLUMN_WIDTH] * len(_WORK_EXPERIENCE_COLUMNS))
    # 보유 전문성(LLM 산출, 부서장/본인 컨펌을 거치지 않은 비객관적 정보)은
    # 다른 옵트인 컬럼과 무엇을 같이 선택하든 항상 맨 마지막 컬럼이 되도록
    # 다른 그룹 뒤에 붙인다.
    if include_expertise:
        columns.extend(_EXPERTISE_COLUMNS)
        widths.extend([_EXPERTISE_COLUMN_WIDTH] * len(_EXPERTISE_COLUMNS))

    wb = Workbook()
    ws = wb.active
    ws.title = '연구원 프로필'

    header_font = Font(name=_FONT_NAME, size=_FONT_SIZE, bold=True)
    body_font = Font(name=_FONT_NAME, size=_FONT_SIZE, bold=False)
    wrap_center = Alignment(wrap_text=True, vertical='center', horizontal='center')
    wrap_left = Alignment(wrap_text=True, vertical='center', horizontal='left')

    for col_idx, (header, _fn) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.border = _BORDER
        cell.alignment = wrap_center

    for row_idx, rid in enumerate(researcher_ids, start=2):
        ctx = _researcher_row_context(rid, tables, permissions, dep_pjt_maps,
                                       eval_excluded_dep_ids, org_code_dep_id_map)
        for col_idx, (_header, fn) in enumerate(columns, start=1):
            value = fn(rid, ctx)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.border = _BORDER
            cell.alignment = wrap_left

    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def default_filename() -> str:
    return f"연구원_프로필_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
