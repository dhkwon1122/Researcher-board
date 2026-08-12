"""
"JOB Market" 탭 핵심 로직 — 과제가 종료된다고 가정했을 때, 그 과제원들이
보유 전문성 기준으로 어떤 다른 과제에 재배치될 수 있는지 추천한다.

흐름:
  1) 종료 예정 과제(과제 단위) 또는 연구원 1명(개인별 검색)을 고른다.
     과제 단위면 jd_reconciliation.get_project_members()로 실제 배정 인원을
     가져오고(과제 직무/대상자 검증과 동일한 org_code 정규화 기준 재사용),
     researcher_profile_export.person_base_table()로 명단(사번/성명/부서/과제/
     CL·년차/학력·전공/나이)을 만든다.
  2) 후보 과제 풀 = project_confl_address.csv에 등록된 전체 과제 중 본인이
     지금 속한 과제(자동 제외)와 사용자가 지정한 제외 부서/과제를 뺀 나머지
     (사용자 확정: 제외 부서를 고르면 그 부서의 모든 과제가, 제외 과제를
     고르면 그 과제 하나만 빠짐 — 둘 다 지정 가능).
  3) 각 후보 과제에 대해 서로 독립적인 두 근거로 유사도를 각각 계산한다
     (사용자 확정: 둘 다 보여주되 구분 표기, 데이터 없는 쪽만 개별 제외):
       A) 과제 분석 기반 — project_expertise_analysis.json(컨플루언스 요약,
          process_project_expertise.py)의 텍스트를 임베딩해 이 사람의
          전문성 프로필 임베딩과 비교.
       B) 과제원 전문성 기반 — 그 과제에 현재 배정된 사람들의 보유 전문성
          프로필을 각각 임베딩해, 이 사람과 가장 가까운 한 명(대표값)과
          비교.
     A/B 둘 다 없는 과제는 후보에서 제외한다(비교할 근거가 아예 없으므로).
  4) A/B 중 더 높은 점수로 상위 후보를 추려(임베딩, 결정적) LLM에게 넘기고,
     LLM이 최종 0~3개만 추리며 사유를 쓴다(구조화 출력 강제 — 순위 자체는
     이미 Python이 계산해 둔 후보 중에서만 고르게 해 할루시네이션 방지).
  5) 개인별 검색도 같은 recommend_for_researcher()를 그대로 재사용한다
     (본인이 현재 속한 과제는 항상 자동 제외 — 사용자 확정).

PII: LLM 프롬프트에는 researcher_id/이름을 절대 포함하지 않는다(이 프로젝트
전체 원칙) — 대상자 본인도, B 근거로 쓰는 과제원 대표 인물도 프로필 텍스트만
전달한다.

Source:
  data/processed/project_confl_address.csv     (후보 과제 전체 목록 — dep_name/project_name)
  data/processed/researchers.csv               (과제 배정 인원 — org_code 기준)
  data/processed/education.csv                 (roster의 학력/전공)
  data/processed/project_expertise_analysis.json (근거 A — process_project_expertise.py)
  data/processed/연구원 보유 전문성 분석.json   (근거 B/대상자 프로필)
"""

import io
import json
import os
import re
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

_PIPELINE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'pipeline')
sys.path.insert(0, os.path.abspath(_PIPELINE_DIR))

import llm_client  # noqa: E402
import researcher_fit as fit  # noqa: E402
from services import data_store  # noqa: E402
from services import jd_reconciliation as jd  # noqa: E402
from services import researcher_profile_export as export  # noqa: E402
from services.llm import LLMError  # noqa: E402

TOP_K_CANDIDATES = 8  # 임베딩으로 추려 LLM에 넘길 상위 후보 수(최종 0~3개는 LLM이 고름)
MAX_RECOMMENDATIONS = 3

HISTORY_DIR = os.path.join(data_store.DATA_DIR, 'job_market')


# ─── 검색 이력 저장/조회 ─────────────────────────────────────────────────────

def _safe_stem(text: str) -> str:
    return re.sub(r'[^0-9A-Za-z가-힣_-]', '_', text or '')[:60] or 'search'


def _history_label(report: dict) -> str:
    if report.get('mode') == 'individual':
        roster = report.get('roster') or []
        if len(roster) == 1:
            p = roster[0]
            return f"{p['name']}({p['researcher_id']})"
        names = ', '.join(p['name'] for p in roster[:3])
        more = f' 외 {len(roster) - 3}명' if len(roster) > 3 else ''
        return f'{names}{more}' if roster else '개인별 검색'
    return ', '.join(report.get('project_names') or [])


def save_history(report: dict) -> str:
    os.makedirs(HISTORY_DIR, exist_ok=True)
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    fname = f"{_safe_stem(_history_label(report))}_{ts}.json"
    with open(os.path.join(HISTORY_DIR, fname), 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    return fname


def list_history() -> list:
    if not os.path.isdir(HISTORY_DIR):
        return []
    rows = []
    for fname in os.listdir(HISTORY_DIR):
        if not fname.endswith('.json'):
            continue
        try:
            with open(os.path.join(HISTORY_DIR, fname), encoding='utf-8') as f:
                data = json.load(f)
        except (OSError, json.JSONDecodeError):
            continue
        rows.append({
            'file': fname,
            'run_at': data.get('run_at', ''),
            'mode': data.get('mode', ''),
            'label': _history_label(data),
            'target_count': len(data.get('roster') or []),
            'candidates_considered': data.get('candidates_considered', 0),
        })
    rows.sort(key=lambda r: r['run_at'], reverse=True)
    return rows


def load_history(file_name: str) -> dict | None:
    """file_name은 list_history()가 돌려준 값만 받는다 — basename만 취급하고
    HISTORY_DIR 밖 경로는 거부해 경로 조작을 막는다."""
    safe_name = os.path.basename(file_name or '')
    if not safe_name.endswith('.json'):
        return None
    path = os.path.join(HISTORY_DIR, safe_name)
    if not os.path.isfile(path):
        return None
    with open(path, encoding='utf-8') as f:
        return json.load(f)


# ─── 과제/부서 목록 (드롭다운, 결정적) ───────────────────────────────────────

def list_departments() -> list:
    """드롭다운용 부서(플랫폼/팀) 목록 — project_confl_address.csv의 dep_name."""
    projects_df = data_store.read_processed('project_confl_address')
    if projects_df.empty or 'dep_name' not in projects_df.columns:
        return []
    return sorted({str(v).strip() for v in projects_df['dep_name'] if str(v).strip()})


def list_projects(department=None) -> list:
    """드롭다운용 과제 목록 — project_confl_address.csv의 project_name.
    department가 있으면 그 dep_name(문자열 1개 또는 복수 선택 시 리스트)인
    과제만."""
    projects_df = data_store.read_processed('project_confl_address')
    if projects_df.empty or 'project_name' not in projects_df.columns:
        return []
    if department:
        depts = {department.strip()} if isinstance(department, str) else {str(d).strip() for d in department}
        projects_df = projects_df[projects_df['dep_name'].astype(str).str.strip().isin(depts)]
    return sorted({str(v).strip() for v in projects_df['project_name'] if str(v).strip()})


# ─── 과제원 명단(roster) — researcher_profile_export의 표시용 컬럼 재사용 ─────

def build_roster(researcher_ids: list) -> list:
    """researcher_id 목록으로 명단(사번/성명/부서/과제/CL·년차/학력·전공/나이)을
    만든다. researcher_profile_export.person_base_table()이 이미 이 7개
    컬럼(및 정확히 같은 "CL/년차"·"학력/전공" 표기)을 계산해 두고 있어 그대로
    재사용한다."""
    base = export.person_base_table(researcher_ids)
    keys = ['researcher_id', 'name', 'department', 'org_code', 'position_year', 'degree_major', 'age']
    return [dict(zip(keys, base.get(rid, [rid, '-', '-', '-', '-', '-', '-']))) for rid in researcher_ids]


# ─── 후보 과제 풀 + 근거 A/B 임베딩 준비 ──────────────────────────────────────

def _dedup_candidate_rows(projects_df) -> list:
    """project_confl_address.csv를 정규화된 project_name(=org_code) 기준으로
    묶어 대표 행 하나씩만 남긴다 — org_code/project_name 표기 형식이 달라도
    (대괄호 태그/띄어쓰기) 같은 과제를 서로 다른 후보로 중복 추천하지 않기
    위함(get_project_members()가 쓰는 것과 동일한 fit.normalize_org_code() 기준)."""
    seen = {}
    for _, row in projects_df.iterrows():
        name = str(row.get('project_name') or '').strip()
        if not name:
            continue
        key = fit.normalize_org_code(name)
        if key not in seen:
            seen[key] = {'project_name': name, 'dep_name': str(row.get('dep_name') or '').strip(), 'norm': key}
    return list(seen.values())


def _normalize_excluded_projects(excluded_org_codes: set) -> set:
    """제외 과제(project_name)를 정규화된 제외 집합으로 만든다. 화면의 '제외할
    부서' 드롭다운은 이 '제외할 과제' 목록을 좁혀 보여주는 캐스케이딩 필터일
    뿐(사용자 확정: 부서 자체는 결과에 아무 영향이 없어야 함) — 실제 후보 풀
    제외는 여기서 만드는 과제 단위 집합만으로 결정된다."""
    return {fit.normalize_org_code(p) for p in excluded_org_codes}


def _build_candidate_pool(excluded_norm: set, expertise_profiles: dict) -> dict:
    """후보 과제별로 근거 A(과제 분석 텍스트)/B(현재 배정 인력 프로필 목록)를
    준비한다. 임베딩은 fit.cached_embed()로 한꺼번에 계산해(텍스트 해시 캐시
    재사용) 후보 수·인원 수가 늘어도 반복 호출을 피한다.

    반환: {project_name(대표 표기): {'dep_name', 'a_text', 'a_embedding'(or None),
                                      'b_members': [(embedding, profile_text), ...]}}
    A/B 둘 다 없는 과제는 이 dict에 아예 포함하지 않는다(비교 근거 없음)."""
    projects_df = data_store.read_processed('project_confl_address')
    if projects_df.empty:
        return {}
    all_rows = _dedup_candidate_rows(projects_df)
    candidates = [row for row in all_rows if row['norm'] not in excluded_norm]
    if not candidates:
        return {}

    # 근거 A 텍스트 준비 (없으면 빈 문자열 — 임베딩 대상에서 제외)
    a_texts_by_project = {}
    for row in candidates:
        entry = jd.read_confluence_summary(row['project_name'])
        text = jd.confluence_context_text(entry)
        if text:
            a_texts_by_project[row['project_name']] = text

    # 근거 B: 후보별 현재 배정 인력의 전문성 프로필 텍스트 (있는 사람만)
    b_texts_by_project = {}
    for row in candidates:
        members = jd.get_project_members(row['project_name'])
        texts = []
        for m in members:
            profile = expertise_profiles.get(m['researcher_id'])
            if profile:
                texts.append(fit.researcher_profile_text(profile))
        if texts:
            b_texts_by_project[row['project_name']] = texts

    # 임베딩은 한 번에 계산(캐시 재사용) — 실패하면 그대로 위로 전파(호출부가 처리)
    all_a_texts = list(a_texts_by_project.values())
    all_b_texts = [t for texts in b_texts_by_project.values() for t in texts]
    a_embeds = fit.cached_embed(all_a_texts) if all_a_texts else None
    b_embeds = fit.cached_embed(all_b_texts) if all_b_texts else None

    a_embed_by_project = {}
    if a_embeds is not None:
        for project_name, vec in zip(a_texts_by_project.keys(), a_embeds):
            a_embed_by_project[project_name] = vec

    b_embed_by_project = {}
    if b_embeds is not None:
        it = iter(b_embeds)
        for project_name, texts in b_texts_by_project.items():
            b_embed_by_project[project_name] = [(next(it), t) for t in texts]

    pool = {}
    for row in candidates:
        project_name = row['project_name']
        a_embedding = a_embed_by_project.get(project_name)
        b_members = b_embed_by_project.get(project_name) or []
        if a_embedding is None and not b_members:
            continue
        pool[project_name] = {
            'dep_name': row['dep_name'],
            'a_text': a_texts_by_project.get(project_name, ''),
            'a_embedding': a_embedding,
            'b_members': b_members,
        }
    return pool


def _cosine(a, b) -> float:
    import numpy as np
    denom = (np.linalg.norm(a) * np.linalg.norm(b)) + 1e-9
    return float(np.dot(a, b) / denom)


def _score_candidates(profile_embedding, candidate_pool: dict) -> list:
    """후보 과제별 A/B 점수를 계산해 결합 점수(둘 중 높은 쪽) 기준 내림차순
    정렬한 리스트로 반환한다."""
    scored = []
    for project_name, cand in candidate_pool.items():
        score_a = _cosine(profile_embedding, cand['a_embedding']) if cand['a_embedding'] is not None else None
        score_b, b_text = None, None
        if cand['b_members']:
            sims = [(_cosine(profile_embedding, emb), text) for emb, text in cand['b_members']]
            score_b, b_text = max(sims, key=lambda x: x[0])
        combined = max(v for v in (score_a, score_b) if v is not None)
        scored.append({
            'project_name': project_name, 'dep_name': cand['dep_name'],
            'score_a': score_a, 'score_b': score_b,
            'a_text': cand['a_text'], 'b_text': b_text or '',
            'combined': combined,
        })
    scored.sort(key=lambda x: -x['combined'])
    return scored


# ─── LLM 최종 판정 (임베딩 상위 후보 중 0~3개만, 사유 포함) ──────────────────

_RECOMMEND_SYSTEM_PROMPT = """# Role
당신은 R&D 인력 재배치를 돕는 "과제 매칭 도우미"입니다. 과제가 종료되어
다른 과제로 옮겨야 하는 연구원 한 명의 보유 전문성과, 여러 후보 과제의
정보를 비교해 실제로 참여 가능성이 높은 과제를 추립니다.

# Goal
주어진 연구원의 전문성 프로필과, 각 후보 과제에 대해 주어지는 근거(A: 과제
자체의 기술 분석 내용, B: 그 과제에 이미 배정된 인력 중 이 사람과 가장 가까운
인력의 전문성 프로필 — 후보마다 A/B 중 하나만 있을 수도, 둘 다 있을 수도
있습니다)를 비교해, 실제로 참여 가능성이 높은 과제를 최대 3개까지 고릅니다.

# Guidelines & Constraints
1. 반드시 입력으로 주어진 후보 과제 중에서만 선택하세요(새로 만들어내지 마세요).
2. 근거가 뚜렷한 순서대로 최대 3개까지만 고르세요. 뚜렷한 연관성이 없으면
   억지로 채우지 말고 빈 리스트를 반환하세요.
3. reason은 왜 이 과제가 맞는지 1~2문장으로 구체적으로 쓰세요. A/B 중 어떤
   근거를 주로 참고했는지 자연스럽게 드러나도록 쓰세요.
4. recommendations가 빈 리스트라면(뚜렷이 맞는 과제가 하나도 없다면), 그래도
   후보 중 그나마 가장 가까운 과제 하나를 골라 closest_non_match에 담으세요.
   reason에는 "추천"이 아니라 왜 이 과제조차 충분히 맞지 않는지(이 연구원의
   전문성과 이 과제가 요구하는 것이 구체적으로 어떻게 다른지)를 1~2문장으로
   쓰세요. recommendations에 1개 이상 있다면 closest_non_match는 반드시
   null로 두세요.
5. 반드시 아래 JSON 형식으로만 출력하세요.

# Output Format (JSON)
{"recommendations": [{"project_name": "후보 목록의 과제명 그대로", "reason": "1~2문장 사유"}],
 "closest_non_match": {"project_name": "후보 목록의 과제명 그대로", "reason": "왜 이것도 충분히 맞지 않는지 1~2문장"} 또는 null}
"""


def _candidate_block(item: dict) -> str:
    a = item['a_text'] or '(과제 분석 데이터 없음)'
    b = item['b_text'] or '(비교 가능한 배정 인력 없음)'
    return f"[{item['project_name']}] 소속: {item['dep_name']}\nA) 과제 분석 기반: {a}\nB) 유사 인력 기반: {b}"


def _judge_recommendations(profile_text: str, shortlist: list) -> tuple:
    """임베딩 상위 후보(shortlist)를 LLM에 보여주고 0~3개만 사유와 함께 고르게
    한다. 반환: (recommendations, closest_non_match). recommendations가
    비어 있으면(뚜렷이 맞는 과제가 없으면) closest_non_match에 "그나마
    가장 가까운 과제 + 왜 이것도 안 맞는지" 근거를 담아 함께 돌려준다 —
    호출부가 "참여 가능한 과제 없음"을 근거 없이 통보하지 않게 한다.
    실패/파싱 오류 시 (빈 리스트, None)으로 안전하게 처리."""
    if not shortlist:
        return [], None
    candidate_block = '\n\n'.join(_candidate_block(c) for c in shortlist)
    prompt = (
        f'[연구원 전문성 프로필]\n{profile_text}\n\n[후보 과제 목록]\n{candidate_block}\n\n'
        '위 연구원이 실제로 참여 가능성이 높은 과제를 최대 3개까지 골라주세요. '
        '뚜렷이 맞는 과제가 없다면 그나마 가장 가까운 과제와 그 사유를 closest_non_match에 담아주세요.'
    )
    raw = llm_client.call_llm(prompt, _RECOMMEND_SYSTEM_PROMPT, temperature=0.0, max_tokens=1200)
    if not raw:
        return [], None
    try:
        parsed = json.loads(llm_client.extract_json(raw))
    except json.JSONDecodeError:
        return [], None

    by_name = {c['project_name']: c for c in shortlist}

    def _attach(r):
        name = str(r.get('project_name') or '').strip()
        cand = by_name.get(name)
        if not cand:
            return None  # 후보 목록에 없는 이름은 할루시네이션으로 간주하고 버림
        return {
            'project_name': name, 'dep_name': cand['dep_name'],
            'reason': str(r.get('reason') or '').strip(),
            'score_a': cand['score_a'], 'score_b': cand['score_b'],
        }

    results = []
    for r in (parsed.get('recommendations') or [])[:MAX_RECOMMENDATIONS]:
        attached = _attach(r)
        if attached:
            results.append(attached)

    closest_non_match = None
    if not results and isinstance(parsed.get('closest_non_match'), dict):
        closest_non_match = _attach(parsed['closest_non_match'])

    return results, closest_non_match


# ─── 전체 실행 ───────────────────────────────────────────────────────────────

def recommend_for_researcher(researcher_id: str, expertise_profiles: dict, candidate_pool: dict) -> dict:
    """한 사람에 대한 추천 결과. candidate_pool은 여러 사람에 걸쳐 재사용
    가능(과제 단위 검색이 배정 인원 전체를 순회할 때 매번 새로 만들지 않도록
    호출부가 미리 만들어 넘긴다).

    반환에 'profile'(연구원 보유 전문성 분석.json의 해당 인물 원본 항목)을
    함께 담는다 — 매칭 과제 사유만으로는 신뢰성이 떨어진다는 사용자 피드백에
    따라, 화면에서 이 사람의 보유 전문성(강점 분야/키워드/역할 등)을 추천
    사유와 나란히 보여주기 위함(components.detail_tabs.llm_summary_block이
    researcher_profile.py와 동일한 형태로 렌더링)."""
    profile = expertise_profiles.get(researcher_id)
    if not profile:
        return {'recommendations': [], 'closest_non_match': None,
                'note': '전문성 분석 데이터가 없어 추천할 수 없습니다.', 'profile': None}
    if not candidate_pool:
        return {'recommendations': [], 'closest_non_match': None,
                'note': '비교 가능한 후보 과제가 없습니다.', 'profile': profile}

    profile_text = fit.researcher_profile_text(profile)
    try:
        profile_embedding = fit.cached_embed([profile_text])[0]
    except LLMError as exc:
        return {'recommendations': [], 'closest_non_match': None,
                'note': f'임베딩 계산 실패: {exc}', 'profile': profile}

    scored = _score_candidates(profile_embedding, candidate_pool)
    shortlist = scored[:TOP_K_CANDIDATES]
    recommendations, closest_non_match = _judge_recommendations(profile_text, shortlist)
    return {'recommendations': recommendations, 'closest_non_match': closest_non_match,
            'note': '', 'profile': profile}


def _own_org_code(researcher_id: str, researchers_df) -> str:
    rows = researchers_df[researchers_df['researcher_id'] == researcher_id]
    if rows.empty:
        return ''
    return str(rows.iloc[0].get('org_code') or '').strip()


def run_project_search(project_names: list, excluded_org_codes: list) -> dict:
    """'종료 예정 과제' 단위 검색 — 복수 선택 가능. 선택한 과제 전체의 배정
    인원(중복 인원은 한 번만)에 대해 추천을 반복하고, 선택한 과제들은 전부
    (본인이 속한 과제뿐 아니라 함께 종료되는 다른 선택 과제들도) 후보에서
    제외한다 — 같이 끝나는 과제로 서로 재추천되지 않도록. 제외는 과제
    단위(excluded_org_codes)로만 결정되고, 화면의 부서 드롭다운은 이 과제
    목록을 좁혀 보여주는 캐스케이딩 필터일 뿐 결과에는 영향을 주지 않는다
    (사용자 확정)."""
    project_names = [p for p in (project_names or []) if p]
    if not project_names:
        return {'error': '종료 예정 과제를 선택해주세요.'}

    researcher_ids = []
    seen_rid = set()
    for project_name in project_names:
        for m in jd.get_project_members(project_name):
            if m['researcher_id'] not in seen_rid:
                seen_rid.add(m['researcher_id'])
                researcher_ids.append(m['researcher_id'])
    if not researcher_ids:
        return {'error': f'선택한 과제({", ".join(project_names)})에 배정된 인원이 없습니다.'}
    roster = build_roster(researcher_ids)

    excluded_norm = _normalize_excluded_projects(set(excluded_org_codes))
    excluded_norm |= {fit.normalize_org_code(p) for p in project_names}  # 종료 예정 과제 전체는 항상 제외

    expertise_profiles = data_store.read_expertise_profiles()
    try:
        candidate_pool = _build_candidate_pool(excluded_norm, expertise_profiles)
    except LLMError as exc:
        return {'error': f'후보 과제 임베딩 계산 실패: {exc}'}

    workers = llm_client.max_concurrency()
    tasks = [(lambda rid=rid: recommend_for_researcher(rid, expertise_profiles, candidate_pool))
              for rid in researcher_ids]
    task_results = llm_client.run_concurrent(tasks, max_workers=workers) if tasks else []

    results = {}
    for rid, (result, error) in zip(researcher_ids, task_results):
        results[rid] = result if error is None and result else {'recommendations': [], 'note': f'처리 중 오류: {error}'}

    report = {
        'mode': 'project', 'project_names': project_names,
        'run_at': datetime.now().isoformat(timespec='seconds'),
        'roster': roster, 'results': results,
        'candidates_considered': len(candidate_pool),
    }
    save_history(report)
    return report


MAX_SEARCH_RESULTS = 20  # 이름/사번 검색창의 후보 표시 개수 상한


def _resolve_researcher_query(query: str, researchers_df) -> list:
    """이름 또는 사번으로 researcher_id 후보를 찾는다(동명이인이면 복수 반환)."""
    if researchers_df.empty:
        return []
    q = str(query or '').strip()
    if not q:
        return []
    qz = q.zfill(8) if q.isdigit() else q
    by_id = researchers_df[researchers_df['researcher_id'] == qz]
    if not by_id.empty:
        return by_id['researcher_id'].tolist()
    by_name = researchers_df[researchers_df['name'].astype(str).str.strip() == q]
    if not by_name.empty:
        return by_name['researcher_id'].tolist()
    contains = researchers_df[researchers_df['name'].astype(str).str.contains(re.escape(q), na=False)]
    return contains['researcher_id'].tolist()


def search_researchers(query: str) -> list:
    """이름/사번 검색창용 — 화면에서 여러 후보 중 골라 추가할 수 있도록,
    동명이인이어도(또는 부분 일치가 여럿이어도) 전부 반환한다(최대
    MAX_SEARCH_RESULTS건). run_individual_search()가 쓰는 정확 일치 대신
    "골라서 추가"하는 흐름이라 여기서는 후보를 좁히지 않는다."""
    researchers_df = data_store.read_processed('researchers')
    candidates = _resolve_researcher_query(query, researchers_df)
    if not candidates:
        return []
    rows = researchers_df[researchers_df['researcher_id'].isin(candidates)]
    return [
        {
            'researcher_id': row['researcher_id'],
            'name': row.get('name', row['researcher_id']),
            'department': row.get('department', ''),
            'org_code': row.get('org_code', ''),
        }
        for _, row in rows.iterrows()
    ][:MAX_SEARCH_RESULTS]


def run_individual_search(researcher_ids: list, excluded_org_codes: list) -> dict:
    """개인별 검색 — 화면에서 미리 골라 확정한 researcher_id 목록(1명 또는
    여러 명)에 대해 각자의 후보 과제를 추천한다. 각자 현재 속한 과제는
    항상 자동으로 제외한다(사람마다 다를 수 있어 개별 계산). 제외는 과제
    단위(excluded_org_codes)로만 결정되고, 화면의 부서 드롭다운은 과제
    목록을 좁혀 보여주는 캐스케이딩 필터일 뿐 결과에는 영향을 주지 않는다
    (사용자 확정).

    같은 과제(정규화된 org_code) 소속 사람이 여럿이면 그들의 제외 조건이
    동일하므로 후보 풀을 그룹당 한 번만 만들어 재사용한다(임베딩 반복
    계산 방지)."""
    researcher_ids = [r for r in dict.fromkeys(researcher_ids or []) if r]  # 순서 유지 중복 제거
    if not researcher_ids:
        return {'error': '검색할 대상자를 추가해주세요.'}

    researchers_df = data_store.read_processed('researchers')
    roster = build_roster(researcher_ids)

    base_excluded_norm = _normalize_excluded_projects(set(excluded_org_codes))
    expertise_profiles = data_store.read_expertise_profiles()

    own_norm_by_rid = {}
    for rid in researcher_ids:
        own_org_code = _own_org_code(rid, researchers_df)
        own_norm_by_rid[rid] = fit.normalize_org_code(own_org_code) if own_org_code else ''

    pool_by_own_org = {}
    try:
        for own_norm in set(own_norm_by_rid.values()):
            excluded_norm = base_excluded_norm | ({own_norm} if own_norm else set())
            pool_by_own_org[own_norm] = _build_candidate_pool(excluded_norm, expertise_profiles)
    except LLMError as exc:
        return {'error': f'후보 과제 임베딩 계산 실패: {exc}'}

    workers = llm_client.max_concurrency()
    tasks = [
        (lambda rid=rid: recommend_for_researcher(rid, expertise_profiles, pool_by_own_org[own_norm_by_rid[rid]]))
        for rid in researcher_ids
    ]
    task_results = llm_client.run_concurrent(tasks, max_workers=workers) if tasks else []

    results = {}
    for rid, (result, error) in zip(researcher_ids, task_results):
        results[rid] = result if error is None and result else {'recommendations': [], 'note': f'처리 중 오류: {error}'}

    report = {
        'mode': 'individual', 'researcher_ids': researcher_ids,
        'run_at': datetime.now().isoformat(timespec='seconds'),
        'roster': roster, 'results': results,
        'candidates_considered': max((len(p) for p in pool_by_own_org.values()), default=0),
    }
    save_history(report)
    return report


# ─── 검색 결과 엑셀 다운로드 ─────────────────────────────────────────────────

def _pct(score) -> str:
    return f'{round(score * 100)}%' if score is not None else '데이터없음'


def build_result_workbook(result: dict) -> bytes:
    """JOB Market 검색 결과(run_project_search/run_individual_search/
    load_history()가 반환하는 report 형태)를 엑셀로 내보낸다. "재배치가
    가능한 연구원의 사번을 첫 번째 컬럼, 결과를 두 번째 컬럼에" 반영해
    달라는 요청대로, 추천이 1건 이상인(재배치 가능한) 연구원만 포함하고
    추천 과제명/부서/A·B 점수/사유를 한 셀에 번호 매겨 줄바꿈 나열한다."""
    roster = result.get('roster') or []
    results = result.get('results') or {}
    ordered_rids = [p['researcher_id'] for p in roster if p.get('researcher_id') in results]
    ordered_rids += [rid for rid in results if rid not in ordered_rids]

    rows = []
    for rid in ordered_rids:
        recs = (results.get(rid) or {}).get('recommendations') or []
        if not recs:
            continue
        blocks = []
        for i, r in enumerate(recs, start=1):
            line = (f"{i}. {r.get('project_name', '')} ({r.get('dep_name', '')}) "
                    f"- A: {_pct(r.get('score_a'))}, B: {_pct(r.get('score_b'))}")
            reason = (r.get('reason') or '').strip()
            if reason:
                line += f'\n   사유: {reason}'
            blocks.append(line)
        rows.append((rid, '\n\n'.join(blocks)))

    wb = Workbook()
    ws = wb.active
    ws.title = 'JOB Market 결과'

    font_name, font_size = '바탕체', 11
    header_font = Font(name=font_name, size=font_size, bold=True)
    body_font = Font(name=font_name, size=font_size, bold=False)
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_center = Alignment(wrap_text=True, vertical='center', horizontal='center')
    wrap_left = Alignment(wrap_text=True, vertical='top', horizontal='left')

    for col_idx, header in enumerate(['사번', '결과'], start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.border = border
        cell.alignment = wrap_center

    for row_idx, (rid, text) in enumerate(rows, start=2):
        id_cell = ws.cell(row=row_idx, column=1, value=rid)
        id_cell.font = body_font
        id_cell.border = border
        id_cell.alignment = wrap_center

        result_cell = ws.cell(row=row_idx, column=2, value=text)
        result_cell.font = body_font
        result_cell.border = border
        result_cell.alignment = wrap_left

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def result_default_filename() -> str:
    return f"JOB_Market_결과_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
