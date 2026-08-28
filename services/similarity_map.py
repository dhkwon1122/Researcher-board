"""
연구원 전문성 임베딩(BGE-M3) 기반 2D 유사도 지도 계산.

pipeline/process_researcher_expertise.py(연구원 보유 전문성 분석.json)와
pipeline/process_researcher_similarity.py가 채워둔 embedding_cache.json
(텍스트 해시 → 1024차원 BGE-M3 벡터)을 읽어, UMAP으로 2D 좌표로 투영한다.

이 서비스는 새로 임베딩을 계산하지 않는다(대시보드 페이지 로드 시 BGE-M3 서버를
호출하지 않음 — 캐시에 없는 연구원은 그냥 지도에서 제외된다). 지도에 모든
연구원이 나오게 하려면 먼저 pipeline/process_researcher_similarity.py를
실행해 embedding_cache.json을 채워야 한다.
"""

import io
import json
import os
from collections import Counter
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from pipeline.rd_specialist_markdown import build_org_tree, read_team_refer
from pipeline.researcher_fit import _text_hash, researcher_profile_text
from services import researcher_profile_export as export
from services.data_store import DATA_DIR, read_expertise_profiles, read_processed, read_similar_researchers


def _cluster_label(rows: list) -> str:
    """클러스터 구성원들의 strength_fields 중 가장 흔한 상위 2개를 조합해
    자동으로 영역 이름을 만든다. 강점 분야 데이터가 전혀 없으면 '미분류 영역'."""
    counter = Counter()
    for r in rows:
        counter.update(r.get('strength_fields') or [])
    if not counter:
        return '미분류 영역'
    return ' · '.join(name for name, _ in counter.most_common(2))


def compute_clusters(df: pd.DataFrame, min_cluster_size: int = 3) -> pd.DataFrame:
    """2D UMAP 좌표(x, y) 위에서 HDBSCAN으로 밀도 기반 클러스터를 찾고, 각
    클러스터에 구성원의 강점 분야 기반 자동 이름(cluster_label)을 붙인다.
    노이즈(어느 클러스터에도 속하지 않는 점, HDBSCAN 라벨 -1)는 cluster=-1,
    cluster_label=''로 남는다 — 화면에서 경계/이름 없이 점만 표시된다.

    좌표 축척(UMAP 출력 범위)이 데이터마다 달라 거리 기반 eps 튜닝이 번거로운
    DBSCAN 대신, 최소 군집 크기만 정하면 되는 HDBSCAN을 쓴다."""
    from sklearn.cluster import HDBSCAN

    if df.empty or len(df) < min_cluster_size:
        df = df.copy()
        df['cluster'] = -1
        df['cluster_label'] = ''
        return df

    clusterer = HDBSCAN(min_cluster_size=min_cluster_size, copy=False)
    labels = clusterer.fit_predict(df[['x', 'y']].to_numpy())

    df = df.copy()
    df['cluster'] = labels

    cluster_labels = {}
    for cid in set(labels):
        if cid == -1:
            continue
        members = df[df['cluster'] == cid].to_dict('records')
        cluster_labels[cid] = _cluster_label(members)
    df['cluster_label'] = df['cluster'].map(cluster_labels).fillna('')
    return df


def compute_medium_clusters(df: pd.DataFrame, min_cluster_size: int = 2) -> pd.DataFrame:
    """1차(소규모) 클러스터 중심좌표를 다시 HDBSCAN으로 묶어, 여러 소규모
    클러스터를 포괄하는 2차(중규모) 클러스터를 만든다. 노이즈 포인트(cluster=-1)와
    다른 소규모 클러스터와 묶이지 못한(고립) 소규모 클러스터는 medium_cluster=-1로
    남는다 — 화면에서는 고립된 소규모 클러스터를 자기 자신만의 상위 경계처럼
    항상 표시해 처리한다(호출부 책임)."""
    df = df.copy()
    small_ids = sorted(c for c in df['cluster'].unique() if c != -1)
    if len(small_ids) < min_cluster_size:
        df['medium_cluster'] = -1
        df['medium_cluster_label'] = ''
        return df

    from sklearn.cluster import HDBSCAN

    centroids = np.array([
        df.loc[df['cluster'] == cid, ['x', 'y']].mean().to_numpy()
        for cid in small_ids
    ])
    clusterer = HDBSCAN(min_cluster_size=min_cluster_size, copy=False)
    medium_labels = clusterer.fit_predict(centroids)
    small_to_medium = dict(zip(small_ids, medium_labels))

    df['medium_cluster'] = df['cluster'].map(small_to_medium).fillna(-1).astype(int)
    df.loc[df['cluster'] == -1, 'medium_cluster'] = -1

    medium_cluster_labels = {}
    for mid in set(medium_labels):
        if mid == -1:
            continue
        members = df[df['medium_cluster'] == mid].to_dict('records')
        medium_cluster_labels[mid] = _cluster_label(members)
    df['medium_cluster_label'] = df['medium_cluster'].map(medium_cluster_labels).fillna('')
    return df


def load_similarity_map(n_neighbors: int = 15, random_state: int = 42, min_cluster_size: int = 3) -> tuple:
    """연구원 전문성 임베딩을 UMAP으로 2D에 투영하고, HDBSCAN으로 유사 전문성
    영역을 클러스터링해 자동 이름을 붙인 DataFrame과, 임베딩 캐시가 없어
    지도에서 빠진 연구원 수를 반환한다.

    반환: (df, missing_count)
      df 컬럼: researcher_id, name, department, org_code, e_support('E'/'R'),
               strength_fields(list), strength_keywords(list), x, y, cluster,
               cluster_label, medium_cluster, medium_cluster_label
      필요 입력 파일이 없거나(연구원 보유 전문성 분석.json/embedding_cache.json)
      좌표를 계산할 만큼 표본이 충분치 않으면(3명 미만) x/y 없는 빈 DataFrame을
      반환한다 — 호출부가 이를 보고 안내 메시지를 보여줄 수 있다."""
    expertise_path = os.path.join(DATA_DIR, '연구원 보유 전문성 분석.json')
    cache_path = os.path.join(DATA_DIR, 'embedding_cache.json')

    if not os.path.exists(expertise_path) or not os.path.exists(cache_path):
        return pd.DataFrame(), 0

    with open(expertise_path, encoding='utf-8') as f:
        profiles = json.load(f)
    with open(cache_path, encoding='utf-8') as f:
        embed_cache = json.load(f)

    researchers_df = read_processed('researchers')
    name_map, dept_map, org_map = {}, {}, {}
    if not researchers_df.empty:
        indexed = researchers_df.set_index('researcher_id')
        name_map = indexed['name'].to_dict()
        dept_map = indexed['department'].to_dict()
        org_map = indexed['org_code'].to_dict()

    tech_ownership_df = read_processed('tech_ownership')
    e_support_map = {}
    if not tech_ownership_df.empty and 'E_support' in tech_ownership_df.columns:
        e_support_map = tech_ownership_df.set_index('researcher_id')['E_support'].to_dict()

    rows, vectors = [], []
    missing = 0
    for item in profiles:
        rid = item.get('researcher_id', '')
        vec = embed_cache.get(_text_hash(researcher_profile_text(item)))
        if vec is None:
            missing += 1
            continue
        vectors.append(vec)
        rows.append({
            'researcher_id': rid,
            'name': name_map.get(rid, ''),
            'department': dept_map.get(rid, '') or '미분류',
            'org_code': org_map.get(rid, '') or '미분류',
            'e_support': e_support_map.get(rid, ''),
            'strength_fields': item.get('strength_fields') or [],
            'strength_keywords': item.get('strength_keywords') or [],
        })

    if len(rows) < 3:
        return pd.DataFrame(rows), missing

    import umap  # 무거운 의존성(numba JIT)이라 실제로 지도를 계산할 때만 임포트

    x = np.array(vectors, dtype=np.float32)
    safe_n_neighbors = max(2, min(n_neighbors, len(rows) - 1))
    reducer = umap.UMAP(n_neighbors=safe_n_neighbors, random_state=random_state)
    coords = reducer.fit_transform(x)

    df = pd.DataFrame(rows)
    df['x'] = coords[:, 0]
    df['y'] = coords[:, 1]
    df = compute_clusters(df, min_cluster_size=min_cluster_size)
    df = compute_medium_clusters(df)
    return df, missing


# ─── 관계 그래프(옵시디언 방식 노드-링크) ─────────────────────────────────────
# UMAP 좌표와는 무관하게, 이미 배치로 판정해 둔 연구원↔연구원 유사도
# (researcher_similarity.json, pipeline/process_researcher_similarity.py)를
# 그대로 엣지로 써서 dash_cytoscape용 elements를 만든다. 점(노드) 위치는
# 임베딩 거리가 아니라 cytoscape의 힘-기반(force-directed) 레이아웃이
# 엣지 연결 관계만으로 그때그때 계산한다.

_GRAPH_PALETTE = [
    '#8ecae6', '#f4a6c6', '#f7dd72', '#95d5b2', '#c9a8f5',
    '#f8b88b', '#7fdbda', '#b8c4e0', '#ffb4a2', '#cdb4db',
]

_LEVEL_ORDER = {'상': 3, '중': 2, '하': 1}


def _dept_class(dept: str, dept_index: dict) -> str:
    idx = dept_index.setdefault(dept, len(dept_index) % len(_GRAPH_PALETTE))
    return f'dept-{idx}'


def build_similarity_graph_elements() -> list:
    """researcher_similarity.json을 dash_cytoscape elements(노드+엣지)로
    변환한다. 노드=연구원(researchers.csv 이름/부서), 엣지=유사도가 있는
    쌍 — 각 연구원의 top-K 목록은 방향성이 있어(A의 목록에 B가 있어도 B의
    목록엔 A가 없을 수 있음) A-B/B-A를 같은 쌍으로 취급해 한 번만
    엣지로 만든다(process_researcher_similarity.py의 _pair_key()와 동일한
    정렬 기준). 같은 쌍이 양쪽에서 서로 다른 score로 나오면(비대칭 판정)
    더 높은 쪽을 채택한다. researcher_similarity.json이 없으면 빈 리스트."""
    similarity_map = read_similar_researchers()
    if not similarity_map:
        return []

    researchers_df = read_processed('researchers')
    name_map, dept_map = {}, {}
    if not researchers_df.empty:
        indexed = researchers_df.set_index('researcher_id')
        name_map = indexed['name'].to_dict()
        dept_map = indexed['department'].to_dict()

    edges_by_key: dict = {}
    node_ids: set = set()
    for rid, entry in similarity_map.items():
        node_ids.add(rid)
        for s in entry.get('similar') or []:
            other = s.get('researcher_id', '')
            if not other or other == rid:
                continue
            node_ids.add(other)
            key = tuple(sorted([rid, other]))
            score = s.get('score')
            existing = edges_by_key.get(key)
            if existing is None or (score is not None and (existing.get('score') is None or score > existing['score'])):
                edges_by_key[key] = {'level': s.get('level', ''), 'score': score, 'evidence': s.get('evidence', '')}

    dept_index: dict = {}
    elements = []
    for rid in sorted(node_ids):
        dept = dept_map.get(rid, '') or '미분류'
        elements.append({
            'data': {
                'id': rid,
                'label': name_map.get(rid, rid),
                'department': dept,
            },
            'classes': _dept_class(dept, dept_index),
        })
    for (a, b), edge in edges_by_key.items():
        evidence = edge['evidence']
        evidence_text = '; '.join(evidence) if isinstance(evidence, list) else str(evidence or '')
        level = edge['level'] if edge['level'] in _LEVEL_ORDER else '하'
        elements.append({
            'data': {
                'source': a, 'target': b,
                'level': edge['level'], 'score': edge['score'], 'evidence': evidence_text,
            },
            'classes': f'level-{level}',
        })
    return elements


def similarity_graph_department_classes() -> list:
    """build_similarity_graph_elements()가 붙인 dept-N 클래스에 맞는
    cytoscape 스타일시트 셀렉터를 순서대로 반환(팔레트 순환)."""
    return [
        {'selector': f'.dept-{i}', 'style': {'background-color': color}}
        for i, color in enumerate(_GRAPH_PALETTE)
    ]


# ─── 조직도 기반 엑셀 다운로드(보유 전문성 + 유사 연구원 명단) ────────────────
# "연구원"/"연구원↔연구원" 탭 아래 별도 패널(pages/researcher_similarity_map.py)
# 에서 쓴다 — 개인별 검색 또는 조직도(team_refer.csv) 부서 단위(하위부서 포함
# 옵션)로 대상을 고르면, 그 사람들의 보유 전문성(LLM 4개 필드)과 저장된 유사
# 연구원 명단 전체(researcher_similarity.json)를 한 줄씩 엑셀로 내보낸다.

def _org_tree() -> list:
    return build_org_tree(read_team_refer(DATA_DIR))


def org_tree_options() -> list:
    """부서 선택 드롭다운 옵션 — 조직도를 들여쓰기로 평탄화해 계층이 눈에
    보이도록 한다. value는 dep_id — dep_id가 없는 노드는 하위 연구원을
    org_code로 특정할 수 없으므로 선택지에서 뺀다(자식은 계속 순회).
    라벨은 pjt_part_name만 사용한다(rd_specialist_markdown.org_tree_html과
    동일한 규칙 — dep_name은 트리 라벨에 관여하지 않는다)."""
    options = []

    def _walk(nodes, depth):
        for node in nodes:
            dep_id = (node.get('dep_id') or '').strip()
            if dep_id:
                indent = '　' * depth
                head = node.get('pjt_part_name') or ''
                who = ' '.join(v for v in (node.get('assignment_name'), node.get('name')) if v)
                label = f'{indent}{head}' + (f' ({who})' if who else '')
                options.append({'label': label, 'value': dep_id})
            _walk(node.get('children') or [], depth + 1)

    _walk(_org_tree(), 0)
    return options


def _index_org_tree_by_dep_id(nodes: list, out: dict) -> dict:
    for node in nodes:
        dep_id = (node.get('dep_id') or '').strip()
        if dep_id:
            out[dep_id] = node
        _index_org_tree_by_dep_id(node.get('children') or [], out)
    return out


def _collect_org_codes(node: dict, include_children: bool) -> set:
    codes = {(node.get('org_name_wd') or '').strip()} - {''}
    if include_children:
        for child in node.get('children') or []:
            codes |= _collect_org_codes(child, True)
    return codes


def researchers_under_departments(dep_ids: list, include_children: bool) -> list:
    """선택된 부서(dep_id, 복수)의 org_code(들)에 속한 연구원 researcher_id
    목록 — include_children이면 선택된 각 부서의 하위 부서까지(조직도 트리
    기준) 전부 포함해 합친다. 매칭되는 연구원이 없으면 빈 리스트."""
    if not dep_ids:
        return []
    by_id = _index_org_tree_by_dep_id(_org_tree(), {})
    org_codes = set()
    for dep_id in dep_ids:
        node = by_id.get(dep_id)
        if node:
            org_codes |= _collect_org_codes(node, include_children)
    if not org_codes:
        return []
    researchers_df = read_processed('researchers')
    if researchers_df.empty or 'org_code' not in researchers_df.columns:
        return []
    matched = researchers_df[researchers_df['org_code'].isin(org_codes)]
    return matched['researcher_id'].tolist()


# ─── "연구원 명단" 화면 부서/과제·파트 검색 필터(pages/researcher_list.py) ──────
# team_refer의 dep_name/pjt_part_name은 조직도 트리 구조(dep_id/upper_dep_id)와
# 무관한 평면(flat) 태그다 — 옵션 목록은 team_refer 원본 행을 그대로 groupby해
# 만들고, 실제 연구원 필터링은 항상 org_name_wd(=researchers.csv의 org_code)
# 매칭을 거친다(라벨 문자열과 researchers.csv 값이 다를 수 있어 라벨로 직접
# 비교하지 않는다 — pages/researcher_list.py의 기존 _project_options() 주석
# 참고: 서로 다른 원천의 표기가 항상 일치한다는 보장이 없다).

def _labels_sorted_by_dep_code(rows: list, label_key: str) -> list:
    """rows에서 label_key(dep_name 또는 pjt_part_name) 고유값을, 그 값을 가진
    행들 중 가장 앞선(가장 작은) dep_code 기준 오름차순으로 정렬해 반환한다
    — 같은 이름이 여러 행에 걸쳐 있어도(예: 여러 파트가 같은 dep_name을
    공유) 조직도 상 가장 먼저 나오는 위치로 정렬 순서를 정한다."""
    best_code: dict[str, str] = {}
    for r in rows:
        name = (r.get(label_key) or '').strip()
        if not name:
            continue
        code = (r.get('dep_code') or '').strip()
        if name not in best_code or code < best_code[name]:
            best_code[name] = code
    return sorted(best_code, key=lambda n: best_code[n])


def department_filter_options() -> list:
    """'부서' 드롭다운 옵션 — team_refer의 dep_name 고유값, 조직코드(dep_code)
    오름차순 정렬(사용자 확정)."""
    names = _labels_sorted_by_dep_code(read_team_refer(DATA_DIR), 'dep_name')
    return [{'label': n, 'value': n} for n in names]


def pjt_part_filter_options(dep_names=None) -> list:
    """'과제/파트' 드롭다운 옵션 — team_refer의 pjt_part_name 고유값,
    조직코드(dep_code) 오름차순 정렬(사용자 확정). dep_names를 지정하면 그
    부서(dep_name)에 속한 행만 남긴다(부서 선택 시 캐스케이딩,
    pages/researcher_list.py의 update_project_options 콜백)."""
    rows = read_team_refer(DATA_DIR)
    if dep_names:
        wanted = {dep_names} if isinstance(dep_names, str) else set(dep_names)
        rows = [r for r in rows if (r.get('dep_name') or '').strip() in wanted]
    names = _labels_sorted_by_dep_code(rows, 'pjt_part_name')
    return [{'label': n, 'value': n} for n in names]


def org_codes_for_dep_names(dep_names, period: tuple | None = None) -> set:
    """선택된 dep_name(들)에 해당하는 team_refer 행들의 org_name_wd 집합.
    period=(시작일, 종료일)을 주면 그 기간 기준(그 기간 안 dep_id별 최신
    스냅샷)으로 매칭한다(2026-08-29 추가, pages/researcher_list.py의
    '누적기준 + 기간 지정' 조회) — 생략하면 기존처럼 오늘 기준."""
    if not dep_names:
        return set()
    wanted = {dep_names} if isinstance(dep_names, str) else set(dep_names)
    return {
        (r.get('org_name_wd') or '').strip()
        for r in read_team_refer(DATA_DIR, period=period)
        if (r.get('dep_name') or '').strip() in wanted
    } - {''}


def dep_name_for_org_code(org_code: str) -> str:
    """researchers.csv의 org_code 하나를 받아, 그 org_code와 매칭되는
    team_refer 행의 dep_name을 반환한다(없으면 빈 문자열) — 연구원 개별
    프로필 화면의 '부서' 드롭다운 기본 선택값을 구할 때 쓴다(연구원 명단의
    '부서' 필터와 동일한 team_refer 기준으로 맞추기 위함)."""
    org_code = (org_code or '').strip()
    if not org_code:
        return ''
    for r in read_team_refer(DATA_DIR):
        if (r.get('org_name_wd') or '').strip() == org_code:
            return (r.get('dep_name') or '').strip()
    return ''


def org_codes_for_pjt_part_names(pjt_part_names, period: tuple | None = None) -> set:
    """선택된 pjt_part_name(들)에 해당하는 team_refer 행들의 org_name_wd 집합.
    period는 org_codes_for_dep_names()와 동일(2026-08-29 추가)."""
    if not pjt_part_names:
        return set()
    wanted = {pjt_part_names} if isinstance(pjt_part_names, str) else set(pjt_part_names)
    return {
        (r.get('org_name_wd') or '').strip()
        for r in read_team_refer(DATA_DIR, period=period)
        if (r.get('pjt_part_name') or '').strip() in wanted
    } - {''}


def title_by_researcher_id(period: tuple | None = None) -> dict:
    """researcher_id(조직 단위 책임자의 사번) → assignment_name(직책) 매핑 —
    team_refer 행 중 조직장급만 사번이 채워져 있다(process_team_refer.py
    참고, 나머지는 매핑이 없어 이 dict에 키 자체가 없다). dep_id별 최신
    (또는 period 지정 시 그 기간 기준 최신) 스냅샷만 쓰는 read_team_refer()를
    거쳐, researcher_id가 있는 행만 추린다(2026-08-29 추가 — 이전엔
    pages/researcher_list.py가 read_processed('team_refer')로 누적된
    원본 테이블 전체를 그대로 dict(zip(...))해, "최신"을 명시적으로
    판정하지 않고 CSV에 저장된 행 순서에 우연히 의존하고 있었다 —
    read_team_refer()로 바꾸면 이 문제도 함께 해결되고, deleted='Y'
    처리된 조직의 리더도 더 이상 매핑에 남지 않는다)."""
    mapping: dict = {}
    for r in read_team_refer(DATA_DIR, period=period):
        rid = (r.get('researcher_id') or '').strip()
        if not rid:
            continue
        mapping.setdefault(rid, (r.get('assignment_name') or '').strip())
    return mapping


def org_code_label_maps(period: tuple | None = None) -> tuple[dict, dict]:
    """researchers.csv의 org_code(=team_refer의 org_name_wd) 전체를 한 번에
    dep_name/pjt_part_name으로 매핑하는 dict 2개(org_code → dep_name,
    org_code → pjt_part_name)를 team_refer 한 번 순회로 만든다.
    dep_name_for_org_code()처럼 한 건씩 team_refer 전체를 매번 다시 훑는
    방식은 명단처럼 수백 행을 한꺼번에 매핑할 때(수백 × team_refer 전체
    스캔) 느리므로, 이 dict를 한 번만 만들어 재사용하는 용도(연구원 명단
    화면 참고). 같은 org_code가 여러 team_refer 행에 걸쳐 있으면(정상
    데이터에서는 드묾) dep_name_for_org_code()와 동일하게 먼저 나온 값을
    쓴다. 매핑이 없는 org_code는 두 dict 어디에도 키가 생기지 않는다 —
    호출부가 `.get(org_code, 원본값)`으로 폴백을 직접 처리한다.

    period=(시작일, 종료일)을 주면 그 기간 기준(그 기간 안 dep_id별 최신
    스냅샷)으로 매핑한다(2026-08-29 추가) — 생략하면 기존처럼 오늘 기준."""
    dep_map: dict = {}
    pjt_map: dict = {}
    for r in read_team_refer(DATA_DIR, period=period):
        org_code = (r.get('org_name_wd') or '').strip()
        if not org_code:
            continue
        dep_map.setdefault(org_code, (r.get('dep_name') or '').strip())
        pjt_map.setdefault(org_code, (r.get('pjt_part_name') or '').strip())
    return dep_map, pjt_map


def individual_search_options() -> list:
    """개인별 검색 드롭다운 옵션 — "이름 [부서] (사번)" 형식(동명이인 구분,
    pages/researcher_profile.py 검색 드롭다운과 동일한 표기 규칙)."""
    researchers_df = read_processed('researchers')
    if researchers_df.empty:
        return []
    options = []
    for _, row in researchers_df.sort_values(['department', 'name']).iterrows():
        dept = str(row.get('department', '') or '').strip()
        dept_suffix = f' [{dept}]' if dept else ''
        options.append({
            'label': f"{row.get('name', '')}{dept_suffix} ({row['researcher_id']})",
            'value': row['researcher_id'],
        })
    return options


# ── 엑셀: 사번/성명/부서/보유전문성(4개)/유사 연구원 명단 ──────────────────────

_RESULT_FONT_NAME = '바탕체'
_RESULT_FONT_SIZE = 11
_RESULT_THIN = Side(style='thin', color='000000')
_RESULT_BORDER = Border(left=_RESULT_THIN, right=_RESULT_THIN, top=_RESULT_THIN, bottom=_RESULT_THIN)

_SIMILARITY_COLUMNS_HEADER = [
    '사번', '성명', '부서',
    '보유전문성(강점 분야)', '보유전문성(강점 키워드)',
    '보유전문성(주요 역할·책임)', '보유전문성(전문지식 및 역량)',
    '유사 연구원 명단',
]


def _similar_list_lines(rid: str, similarity_map: dict, name_map: dict, dept_map: dict) -> str:
    """researcher_similarity.json에 저장된 유사 연구원 전체(화면의 3/5/10
    표시 개수 제한과 무관하게)를 "이름(부서) - 유사도% [판정]" 줄로 나열."""
    similar = (similarity_map.get(rid) or {}).get('similar') or []
    if not similar:
        return '-'
    lines = []
    for s in similar:
        srid = s.get('researcher_id', '')
        name = name_map.get(srid, srid)
        dept = dept_map.get(srid, '') or '-'
        level = s.get('level', '')
        score = s.get('score')
        score_disp = f'{round(score * 100)}%' if score is not None else '데이터없음'
        level_disp = f' [{level}]' if level else ''
        lines.append(f'{name}({dept}) - {score_disp}{level_disp}')
    return '\n'.join(lines)


def build_expertise_similarity_workbook(researcher_ids: list) -> bytes:
    """선택된 연구원들의 보유 전문성(LLM 4개 필드)과 저장된 유사 연구원 명단
    전체를 엑셀로 내보낸다. 개인별 검색 또는 조직도 부서 단위(하위부서 포함
    옵션)로 고른 researcher_ids를 그대로 받는다(중복은 순서 유지하며 제거)."""
    researcher_ids = list(dict.fromkeys(researcher_ids))

    researchers_df = read_processed('researchers')
    name_map, dept_map = {}, {}
    if not researchers_df.empty:
        indexed = researchers_df.set_index('researcher_id')
        name_map = indexed['name'].to_dict()
        dept_map = indexed['department'].to_dict()

    profiles = read_expertise_profiles()
    similarity_map = read_similar_researchers()

    wb = Workbook()
    ws = wb.active
    ws.title = '보유 전문성·유사 연구원'

    header_font = Font(name=_RESULT_FONT_NAME, size=_RESULT_FONT_SIZE, bold=True)
    body_font = Font(name=_RESULT_FONT_NAME, size=_RESULT_FONT_SIZE, bold=False)
    wrap_center = Alignment(wrap_text=True, vertical='center', horizontal='center')
    wrap_left = Alignment(wrap_text=True, vertical='top', horizontal='left')

    for col_idx, header in enumerate(_SIMILARITY_COLUMNS_HEADER, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.border = _RESULT_BORDER
        cell.alignment = wrap_center

    for row_idx, rid in enumerate(researcher_ids, start=2):
        profile = profiles.get(rid)
        values = [
            rid,
            name_map.get(rid, '') or '-',
            dept_map.get(rid, '') or '-',
            export.expertise_field_lines(profile, 'strength_fields'),
            export.expertise_field_lines(profile, 'strength_keywords'),
            export.expertise_field_lines(profile, 'key_responsibilities'),
            export.expertise_field_lines(profile, 'domain_knowledge_skill'),
            _similar_list_lines(rid, similarity_map, name_map, dept_map),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.border = _RESULT_BORDER
            cell.alignment = wrap_center if col_idx <= 3 else wrap_left

    for col_idx, width in enumerate([12, 14, 18, 24, 24, 26, 26, 40], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def similarity_workbook_filename() -> str:
    return f"보유전문성_유사연구원_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"


def build_researcher_mail_html(rid: str) -> str | None:
    """연구원 한 명의 보유 전문성 + 유사 연구원 매칭 결과를 메일 발송용 HTML로
    만든다(pages/researcher_similarity_map.py의 메일 발송 모달이 사용).
    build_html()이 만드는 조직도 리포트 전체가 아니라, 이 사람 카드 2장만
    뽑아 사이드바 없는 단순 페이지(mail_page())로 감싼다 — 앱 밖으로 나가는
    메일이니만큼 다른 사람 정보는 섞이지 않도록. 프로필/메일 링크(target="_top"
    상대경로라 앱 밖 메일 본문에서는 깨짐 — 사용자 확인)는 include_links=False로
    빼고 렌더링한다. 보유 전문성 데이터 자체가 없으면(파이프라인 미실행/
    분석 대상 아님) None."""
    from pipeline.process_researcher_expertise import researcher_card_html
    from pipeline.process_researcher_similarity import researcher_match_card_html
    from pipeline.rd_specialist_markdown import mail_page

    profile = read_expertise_profiles().get(rid)
    if not profile:
        return None

    researchers_df = read_processed('researchers')
    name_map, dept_map, org_map = {}, {}, {}
    if not researchers_df.empty:
        indexed = researchers_df.set_index('researcher_id')
        name_map = indexed['name'].to_dict()
        dept_map = indexed['department'].to_dict()
        org_map = indexed['org_code'].to_dict()
    name = name_map.get(rid, rid)

    sections = [
        '<h2>보유 전문성</h2>',
        researcher_card_html(profile, name_map, include_links=False),
    ]

    similarity_item = read_similar_researchers().get(rid)
    if similarity_item and similarity_item.get('similar'):
        profile_by_id = read_expertise_profiles()
        sections.append('<h2>유사 연구원 매칭</h2>')
        sections.append(researcher_match_card_html(
            similarity_item, name_map, dept_map, org_map, profile_by_id, include_links=False,
        ))

    return mail_page(f'{name}({rid}) 보유 전문성', ''.join(sections))
